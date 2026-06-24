"""
Product Import Generator - KING ERP Import
==========================================
Genereert een KING-importexcel voor nieuwe producten.

Input:  Excel/CSV met minimaal: Artikelnummer, Omschrijving, Merk
Output: KING import Excel klaar voor Excel2King

Gebruik:
    python nieuw_product_generator.py   (via app.py of standalone)
"""

import io
import json
import re
import time
from datetime import datetime
from pathlib import Path
from urllib.parse import urljoin, urlparse

import pypdf

try:
    from ddgs import DDGS as _DDGS
    _DDG_BESCHIKBAAR = True
except ImportError:
    try:
        from duckduckgo_search import DDGS as _DDGS
        _DDG_BESCHIKBAAR = True
    except ImportError:
        _DDG_BESCHIKBAAR = False

import httpx
import openpyxl
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openai import OpenAI
from PIL import Image

import config
from de_taalcodes import genereer_de_taalcodes, DE_KOLOMMEN, DE_VELDEN
from jl_taalcodes import genereer_jl_taalcodes, JL_KOLOMMEN, JL_VELDEN

# ─────────────────────────────────────────────
#  MERKDOMEINEN
# ─────────────────────────────────────────────

def laad_merk_domeinen(pad: Path) -> dict:
    """Laad merk → domein mapping uit JSON. Keys worden lowercase opgeslagen."""
    if not pad or not pad.exists():
        return {}
    try:
        with open(pad, encoding="utf-8") as f:
            data = json.load(f)
        return {k.lower().strip(): v.strip() for k, v in data.items()
                if not k.startswith("_") and v and isinstance(v, str)}
    except Exception:
        return {}


def zoek_leverancier_url(merk: str, omschrijving: str, merk_domeinen: dict) -> str:
    """
    Zoekt de productpagina op het leveranciersdomein via DDG site:-zoekopdracht.
    Retourneert de beste URL of "" als niets bruikbaars gevonden.
    """
    if not _DDG_BESCHIKBAAR or not merk_domeinen:
        return ""

    domein = merk_domeinen.get(merk.lower().strip(), "")
    if not domein:
        return ""

    # Strip maat/inhoud voor schonere query (bijv. "Sikaflex-291 600ml" → "Sikaflex-291")
    omschr_zoek = _strip_variant_suffix(omschrijving).strip()
    query = f"site:{domein} {omschr_zoek}"

    try:
        with _DDGS(verify=False) as ddg:
            resultaten = ddg.text(query, max_results=5)
    except Exception:
        return ""

    for r in resultaten:
        url = r.get("href", "")
        if url and domein in url.lower():
            return url

    return ""


# Doelmap productafbeeldingen — NAS in productie, lokale testmap in testmodus
AFBEELDING_BASE_DIR = (
    config._TEST_DIR / "Productafbeeldingen"
    if config.TEST_MODUS
    else config.AFBEELDING_NAS_PAD
)
AFBEELDING_GROOTTE  = (1600, 1600)
# Minimale afmeting om logo's/iconen te filteren
AFBEELDING_MIN_PX   = 200

# ─────────────────────────────────────────────
#  LEVERANCIERSPAGINA OPHALEN
# ─────────────────────────────────────────────

_PIB_TREFWOORDEN = ["productinformatieblad", "informatieblad", "tds", "technical data", "productblad", "pib"]
_VIB_TREFWOORDEN = ["veiligheidsblad", "safety data", "sds", "msds", "vib"]
_HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}


def _schone_bestandsnaam(tekst: str) -> str:
    """Verwijder tekens die niet in bestandsnamen mogen."""
    return re.sub(r'[\\/:*?"<>|]', "", tekst).strip()


def _download_pdf(pdf_url: str, save_pad: Path) -> bool:
    """Download een PDF naar save_pad. Retourneert True bij succes."""
    try:
        save_pad.parent.mkdir(parents=True, exist_ok=True)
        with httpx.stream("GET", pdf_url, timeout=30, follow_redirects=True,
                          verify=False, headers=_HEADERS) as r:
            r.raise_for_status()
            with open(save_pad, "wb") as f:
                for chunk in r.iter_bytes(chunk_size=8192):
                    f.write(chunk)
        return True
    except Exception:
        return False


def _verwerk_afbeelding(img_bytes: bytes, save_pad: Path) -> bool:
    """
    Schaal afbeelding naar 1600x1600 met witte achtergrond (padding).
    Slaat op als JPEG. Retourneert True bij succes.
    """
    try:
        img = Image.open(io.BytesIO(img_bytes)).convert("RGBA")
        if img.width < AFBEELDING_MIN_PX or img.height < AFBEELDING_MIN_PX:
            return False
        # Schaal naar maximaal 1600x1600 met behoud van verhouding (omhoog én omlaag)
        schaal = min(AFBEELDING_GROOTTE[0] / img.width, AFBEELDING_GROOTTE[1] / img.height)
        nieuw_w = round(img.width * schaal)
        nieuw_h = round(img.height * schaal)
        img = img.resize((nieuw_w, nieuw_h), Image.LANCZOS)
        canvas = Image.new("RGB", AFBEELDING_GROOTTE, (255, 255, 255))
        offset = ((AFBEELDING_GROOTTE[0] - nieuw_w) // 2,
                  (AFBEELDING_GROOTTE[1] - nieuw_h) // 2)
        canvas.paste(img, offset, mask=img.split()[3] if img.mode == "RGBA" else None)
        save_pad.parent.mkdir(parents=True, exist_ok=True)
        canvas.save(save_pad, "JPEG", quality=92)
        return True
    except Exception:
        return False


def _download_afbeeldingen(soup: BeautifulSoup, base_url: str,
                            merk: str, bestandsnaam_basis: str) -> list[str]:
    """
    Zoek productafbeeldingen op de pagina, download en sla op als 1600x1600 JPEG.
    Filtert op alt-tekst of bestandsnaam die merknaam of productnaam bevat.
    Retourneert lijst van lokale paden (strings).
    """
    merk_slug = merk.lower().strip()
    merk_dir  = AFBEELDING_BASE_DIR / merk_slug
    paden: list[str] = []
    gezien: set[str] = set()

    # Zoekwoorden waarop gefilterd wordt (merk + losse woorden uit omschrijving)
    basis_slug   = re.sub(r"[^a-z0-9]", "", bestandsnaam_basis.lower())
    zoekwoorden  = [re.sub(r"[^a-z0-9]", "", merk_slug)]
    for woord in bestandsnaam_basis.lower().split():
        slug = re.sub(r"[^a-z0-9]", "", woord)
        if len(slug) > 2:
            zoekwoorden.append(slug)

    def _is_productafbeelding(src: str, alt: str) -> bool:
        # Alleen bestandsnaam checken (niet domein — dat geeft valse matches)
        bestandsnaam = Path(urlparse(src).path).name
        combined = re.sub(r"[^a-z0-9]", "", (bestandsnaam + " " + alt).lower())
        return any(w in combined for w in zoekwoorden)

    # Verzamel kandidaat-afbeelding-URLs
    kandidaten: list[tuple[str, str]] = []  # (full_url, alt)
    for img in soup.find_all("img"):
        src = img.get("src") or img.get("data-src") or img.get("data-lazy-src") or ""
        if not src:
            continue
        full = urljoin(base_url, src)
        if full in gezien:
            continue
        gezien.add(full)
        if any(x in full.lower() for x in [".svg", ".gif", "data:"]):
            continue
        alt = img.get("alt", "")
        if _is_productafbeelding(full, alt):
            kandidaten.append((full, alt))

    teller = 0
    for img_url, alt in kandidaten:
        if teller >= 6:  # max 6 afbeeldingen per product
            break
        try:
            r = httpx.get(img_url, timeout=15, follow_redirects=True,
                          verify=False, headers=_HEADERS)
            r.raise_for_status()
        except Exception:
            continue

        suffix = ".jpg" if teller == 0 else f" {teller + 1}.jpg"
        save_pad = merk_dir / f"{bestandsnaam_basis}{suffix}"
        if _verwerk_afbeelding(r.content, save_pad):
            paden.append(str(save_pad))
            teller += 1

    return paden


def _kandidaat_bytes_van_pagina(
    page_url: str,
    zoekwoorden: list[str],
) -> tuple[bytes, str] | None:
    """
    Haalt de eerste passende productafbeelding op van een pagina als raw bytes.
    Slaat niets op schijf op. Retourneert (bytes, img_url) of None.
    """
    try:
        resp = httpx.get(page_url, timeout=15, follow_redirects=True,
                         verify=False, headers=_HEADERS)
        resp.raise_for_status()
    except Exception:
        return None

    soup = BeautifulSoup(resp.text, "html.parser")
    gezien: set[str] = set()

    for img in soup.find_all("img"):
        src = img.get("src") or img.get("data-src") or img.get("data-lazy-src") or ""
        if not src:
            continue
        full = urljoin(page_url, src)
        if full in gezien:
            continue
        gezien.add(full)
        if any(x in full.lower() for x in [".svg", ".gif", "data:"]):
            continue
        bestandsnaam = Path(urlparse(full).path).name
        alt = img.get("alt", "")
        combined = re.sub(r"[^a-z0-9]", "", (bestandsnaam + " " + alt).lower())
        if not any(w in combined for w in zoekwoorden):
            continue
        try:
            r = httpx.get(full, timeout=15, follow_redirects=True,
                          verify=False, headers=_HEADERS)
            r.raise_for_status()
            img_obj = Image.open(io.BytesIO(r.content))
            if img_obj.width < AFBEELDING_MIN_PX or img_obj.height < AFBEELDING_MIN_PX:
                continue
            return r.content, full
        except Exception:
            continue

    return None


def _beoordeel_afbeeldingen_vision(
    kandidaten: list[dict],
    client,
    omschrijving: str,
) -> int:
    """
    Vraagt GPT-4o Vision welke kandidaat het beste productfoto is.
    Retourneert 0-based index van de beste kandidaat, of 0 als fallback.
    """
    if len(kandidaten) <= 1:
        return 0

    import base64

    inhoud: list[dict] = [
        {
            "type": "text",
            "text": (
                f"Je beoordeelt {len(kandidaten)} productfoto-kandidaten voor: {omschrijving}.\n"
                "Welke afbeelding toont het meest duidelijk het product zelf (bijv. een verfblik, tube "
                "of verpakking) op een witte of neutrale achtergrond — zonder kleurvlak, zonder "
                "moodboard/lifestyle-foto, en scherp genoeg?\n"
                f"Antwoord met ALLEEN een getal (1 t/m {len(kandidaten)}), "
                "of 0 als geen enkele geschikt is als hoofdproductfoto."
            )
        }
    ]

    for i, k in enumerate(kandidaten, 1):
        try:
            b64 = base64.b64encode(k["bytes"]).decode()
            img = Image.open(io.BytesIO(k["bytes"]))
            fmt = (img.format or "JPEG").lower()
            mime = f"image/{'jpeg' if fmt in ('jpg', 'jpeg') else fmt}"
            inhoud.append({
                "type": "image_url",
                "image_url": {"url": f"data:{mime};base64,{b64}", "detail": "low"}
            })
        except Exception:
            inhoud.append({"type": "text", "text": f"[Afbeelding {i}: niet laadbaar]"})

    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            max_tokens=5,
            messages=[{"role": "user", "content": inhoud}]
        )
        antwoord = response.choices[0].message.content.strip()
        match = re.search(r"\d+", antwoord)
        if not match:
            return 0
        gekozen = int(match.group())
        if gekozen == 0 or gekozen > len(kandidaten):
            return 0
        return gekozen - 1
    except Exception:
        return 0


def haal_leverancier_pagina(url: str, merk: str = "", omschrijving: str = "") -> dict:
    """
    Haalt leverancierspagina op en retourneert:
      tekst     - leesbare paginatekst voor AI-prompt
      pib_pad   - lokaal K-schijf pad van gedownload PIB (of "")
      vib_pad   - lokaal K-schijf pad van gedownload VIB (of "")
    """
    if not url or not url.startswith("http"):
        return {"tekst": "", "pib_pad": "", "vib_pad": ""}

    try:
        resp = httpx.get(url, timeout=15, follow_redirects=True,
                         verify=False, headers=_HEADERS)
        resp.raise_for_status()
    except Exception as e:
        return {"tekst": f"[Pagina ophalen mislukt: {e}]", "pib_pad": "", "vib_pad": ""}

    soup = BeautifulSoup(resp.text, "html.parser")

    for tag in soup(["script", "style", "nav", "footer", "header"]):
        tag.decompose()
    tekst = " ".join(soup.get_text(separator=" ").split())[:4000]

    # PDF-links zoeken
    pib_url = vib_url = ""
    base = f"{urlparse(url).scheme}://{urlparse(url).netloc}"

    for a in soup.find_all("a", href=True):
        href = a["href"].lower()
        combined = href + " " + a.get_text().lower()
        if not pib_url and any(t in combined for t in _PIB_TREFWOORDEN):
            pib_url = urljoin(base, a["href"])
        if not vib_url and any(t in combined for t in _VIB_TREFWOORDEN):
            vib_url = urljoin(base, a["href"])

    # PDFs downloaden naar K:\ structuur (zelfde als FLEX Excel)
    # K:\1. Productinformatiebladen\{Merk}\NL\PIB {Merk} {Omschrijving}-NL.pdf
    pib_pad = vib_pad = ""
    merk = merk or ""
    omschrijving = omschrijving or ""
    merk_clean = _schone_bestandsnaam(merk)
    omschr_stripped = omschrijving.strip()
    if merk and omschr_stripped.lower().startswith(merk.lower()):
        omschr_stripped = omschr_stripped[len(merk):].strip()
    omschr_clean = _schone_bestandsnaam(omschr_stripped).strip(" -")
    merk_dir = config.PIB_BASE_DIR / merk_clean / "NL"

    if pib_url:
        pib_bestand = merk_dir / f"PIB {merk_clean} {omschr_clean}-NL.pdf"
        if _download_pdf(pib_url, pib_bestand):
            pib_pad = str(pib_bestand)

    if vib_url:
        vib_bestand = merk_dir / f"VIB {merk_clean} {omschr_clean}-NL.pdf"
        if _download_pdf(vib_url, vib_bestand):
            vib_pad = str(vib_bestand)

    # Afbeeldingen downloaden en opslaan (map wordt aangemaakt indien nodig)
    afbeelding_basis = f"{merk_clean.title()} {omschr_clean}" if omschr_clean else merk_clean.title()
    afbeeldingen = _download_afbeeldingen(soup, url, merk_clean, afbeelding_basis)

    return {"tekst": tekst, "pib_pad": pib_pad, "vib_pad": vib_pad, "afbeeldingen": afbeeldingen}

# ─────────────────────────────────────────────
#  VARIANT-GROEP DETECTIE
# ─────────────────────────────────────────────

# Variant-suffixen die aan het einde van een omschrijving kunnen staan
_VARIANT_RE = re.compile(
    r"""
    \s*
    \d+[\.,]?\d*          # getal (bijv. 5, 0.5, 2,5)
    \s*
    (?:
        kg|g|gram|
        liter|ltr|l|ml|cl|
        stuks?|stk|st\b|
        m2|m²|m\b|cm|mm|
        pak|blik|bus|emmer|set|doos|rol|tube|fles|spuit|can|
        x\d+              # bijv. 2x500ml
    )
    \b
    (?:\s+\d+[\.,]?\d*\s*(?:kg|g|liter|l|ml|stuks?|stk|pak|blik|bus|emmer|set|doos|rol))?
    """,
    re.IGNORECASE | re.VERBOSE,
)

def _strip_variant_suffix(omschrijving: str) -> str:
    """Geeft de basisnaam terug zonder maat/inhoud/stuks aan het einde."""
    return _VARIANT_RE.sub("", omschrijving).strip().rstrip(",-").strip()


def _gemeenschappelijk_prefix(strings: list[str]) -> str:
    """Langste gemeenschappelijke prefix (karakter-niveau) van een lijst strings."""
    if not strings:
        return ""
    prefix = strings[0]
    for s in strings[1:]:
        while not s.startswith(prefix):
            prefix = prefix[:-1]
            if not prefix:
                return ""
    return prefix


def _gemeenschappelijk_prefix_woorden(beschrijvingen: list[str]) -> str:
    """Langste gemeenschappelijke prefix op woordniveau (case-insensitief)."""
    if not beschrijvingen:
        return ""
    woorden = [b.lower().split() for b in beschrijvingen]
    gemeenschappelijk = []
    for i in range(min(len(w) for w in woorden)):
        if all(w[i] == woorden[0][i] for w in woorden):
            gemeenschappelijk.append(woorden[0][i])
        else:
            break
    # Herstel originele schrijfwijze van het eerste artikel
    origineel = beschrijvingen[0].split()
    return " ".join(origineel[:len(gemeenschappelijk)])


def _bouw_groep(leden: list[dict], groep_id: int) -> list[dict]:
    """Maakt parent + children voor een groep varianten."""
    artnrs    = [str(p.get("Artikelnummer", "")) for p in leden]
    prefix    = _gemeenschappelijk_prefix(artnrs)
    prefix    = re.sub(r'\d+$', '', prefix)   # strip losse cijfers aan het einde (bijv. "5630BNA03" → "5630BNA")
    parent_nr = f"{prefix}0000WEB" if prefix else f"WEBGRP{groep_id:04d}"

    omschrijvingen = [str(p.get("Omschrijving", "")) for p in leden]
    basisnaam      = _gemeenschappelijk_prefix_woorden(omschrijvingen).strip()
    if not basisnaam:
        basisnaam = _strip_variant_suffix(omschrijvingen[0])

    # Bepaal configuratie-dimensies:
    # Zijn er variaties in de basis-omschrijving (na stripping van maat)?
    bases_gestript = [_strip_variant_suffix(o).lower() for o in omschrijvingen]
    heeft_type_var = len(set(bases_gestript)) > 1
    heeft_maat_var = any(_VARIANT_RE.search(o) for o in omschrijvingen)

    if heeft_type_var and heeft_maat_var:
        config_op = "Type,Inhoud"
    elif heeft_type_var:
        config_op = "Type"
    else:
        config_op = "Inhoud"

    eerste     = leden[0]
    parent_rij = {
        **{k: v for k, v in eerste.items()
           if k not in ("Artikelnummer", "Omschrijving", "Eenheid", "EanCode")},
        "Artikelnummer":    parent_nr,
        "Omschrijving":     basisnaam,
        "Eenheid":          "",
        "EanCode":          "",
        "_type":            "parent",
        "_configuratie_op": config_op,
        "_groep_id":        groep_id,
    }
    rijen = [parent_rij]
    for p in leden:
        rijen.append({
            **p,
            "_type":            "child",
            "_parent_code":     parent_nr,
            "_configuratie_op": config_op,
            "_groep_id":        groep_id,
        })
    return rijen


def detecteer_variantgroepen(producten: list[dict]) -> list[dict]:
    """
    Groepeert producten die varianten zijn van hetzelfde product.

    Twee passes:
      1. Groepeer op exacte basisnaam na afstropen van maat/inhoud-suffixen.
         → vangt "Anti-Skid Coarse 3kg" + "Anti-Skid Coarse 25kg"
      2. Merge groepen waarvan de basisnamen een gemeenschappelijk woordprefix
         delen van ≥ 3 woorden.
         → vangt "Anti-Skid Coarse" + "Anti-Skid Medium" + "Anti-Skid Fine"
    """
    from collections import defaultdict

    al_gemarkeerd = [p for p in producten if p.get("_type")]
    te_analyseren  = [p for p in producten if not p.get("_type")]

    # ── Pass 1: groepeer op exacte gestripte basisnaam ──────────────────────
    basis_groepen: dict[str, list[dict]] = defaultdict(list)
    for p in te_analyseren:
        basis = _strip_variant_suffix(str(p.get("Omschrijving", ""))).lower().strip()
        basis_groepen[basis].append(p)

    # ── Pass 2: merge groepen met ≥ 3 gemeenschappelijke woorden ────────────
    # Bouw union-find structuur
    bases   = list(basis_groepen.keys())
    ouder   = {b: b for b in bases}   # union-find

    def vind(x):
        while ouder[x] != x:
            ouder[x] = ouder[ouder[x]]
            x = ouder[x]
        return x

    def unie(a, b):
        ra, rb = vind(a), vind(b)
        if ra != rb:
            # Gebruik de kortste als root (= meest generieke naam)
            ouder[rb] = ra if len(ra) <= len(rb) else rb
            if len(rb) < len(ra):
                ouder[ra] = rb

    for i, basis_a in enumerate(bases):
        woorden_a = basis_a.split()
        for basis_b in bases[i + 1:]:
            woorden_b = basis_b.split()
            gedeeld = 0
            for wa, wb in zip(woorden_a, woorden_b):
                if wa == wb:
                    gedeeld += 1
                else:
                    break
            if gedeeld >= 3:
                unie(basis_a, basis_b)

    # Verzamel super-groepen
    super_groepen: dict[str, list[dict]] = defaultdict(list)
    for basis, leden in basis_groepen.items():
        super_groepen[vind(basis)].extend(leden)

    # ── Resultaat opbouwen ───────────────────────────────────────────────────
    resultaat = list(al_gemarkeerd)
    groep_id  = 0

    for _, leden in super_groepen.items():
        if len(leden) < 2:
            resultaat.extend(leden)
        else:
            groep_id += 1
            resultaat.extend(_bouw_groep(leden, groep_id))

    return resultaat


# ─────────────────────────────────────────────
#  EXCEL KOLOMDEFINITIE (zelfde structuur als FLEX import)
# ─────────────────────────────────────────────

KOLOMMEN = [
    # ── KING Import kolommen ─────────────────────────────────────
    "Artikelnummer",                        # 1
    "Eenheid",                              # 2
    "Zoekcode",                             # 3  ← Merk
    "Omschrijving",                         # 4
    "Opbrengstgroep",                       # 5
    "WebArtikel",                           # 6  = 1
    "TekstOpFactuur",                       # 7  ← Omschrijving (invoer)
    "AfbeeldingKlein",                      # 8
    "AfbeeldingGroot",                      # 9
    "Leveranciernummer",                    # 10 ← handmatig (oranje)
    "Leveranciernaam",                      # 11 ← Merk
    "ArtikelOmschrijvingLeverancier",       # 12 ← Omschrijving
    "ArtikelNummerBijLeverancier",          # 13
    "EanCode",                              # 14
    "VR_ART_Magentotype",                   # 15 = "Simpel" / "Configureerbaar"
    "VR_ART_Zichtbaarheid",                 # 16 = "Catalogus, zoeken" / "Niet individueel zichtbaar"
    "VR_ART_Virtueel_Artikel_SKU",          # 17 = parent code (child) of eigen code (parent)
    "VR_ART_Configuratie_op",              # 18 = "Inhoud,Kleur" voor configureerbare producten
    "VR_ART_Actief_in_shop",               # 19 = 1
    "VR_ART_F-Merk",                       # 20 ← Merk
    "VR_ART_Extra_afbeelding_1",           # 21
    "VR_ART_Extra_afbeelding_2",           # 22
    "VR_ART_Extra_afbeelding_3",           # 23
    "VR_ART_Extra_afbeelding_4",           # 24
    "VR_ART_Extra_afbeelding_5",           # 25
    "VR_ART_Productinformatieblad_NL",     # 26
    "VR_ART_Productveiligheidsblad_NL_A",  # 27
    "VR_ART_Attribuutset_V4",              # 28 ← AI
    "1NT_FOV_NL_TITLE_WEB",               # 29 ← AI
    "1NF_FOV_NL_TITLE_FACTUUR",           # 30 ← AI
    "1NH_FOV_NL_URL",                     # 31 ← AI
    "1N1_FOV_NL_META_DATA_1",             # 32 ← AI
    "1N2_FOV_NL_META_DATA_2",             # 33 ← AI
    "1NL_FOV_NL_LANGE_OMSCHRIJVING",      # 34 ← AI
    # ── Extra controle-kolommen (niet in FLEX, voor review) ──────
    "Attribuutset_Confidence_%",
    "Attribuutset_Label",
    "VR_ART_Webcategorie_ID._V2",
    "Webcategorie_Confidence_%",
]

HANDMATIG    = {"Leveranciernummer", "VerkoopPrijsExBTW", "AdviesPrijsExBTW"}
AI_VELDEN    = {
    "1N1_FOV_NL_META_DATA_1", "1N2_FOV_NL_META_DATA_2",
    "1NL_FOV_NL_LANGE_OMSCHRIJVING", "1NT_FOV_NL_TITLE_WEB",
    "1NF_FOV_NL_TITLE_FACTUUR", "1NH_FOV_NL_URL",
}
MAPPING_VELDEN = {
    "VR_ART_Attribuutset_V4", "Attribuutset_Confidence_%", "Attribuutset_Label",
    "VR_ART_Webcategorie_ID._V2", "Webcategorie_Confidence_%",
}


# ─────────────────────────────────────────────
#  REFERENTIEDATA LADEN
# ─────────────────────────────────────────────

def laad_attribuutset(pad: Path) -> list[dict]:
    """
    Laad attribuutset V4 uit het BRON-bestand.
    Kolom 0 (Keuzelijst) = attribuutset code.
    Kolom 6 (Kolom3)     = taxonomy-omschrijving als label.
    """
    if not pad.exists():
        return []
    df = pd.read_excel(pad, header=0)
    SKIP = {"nan", "-", "keuzelijst", ""}
    result = []
    for _, row in df.iterrows():
        code = str(row.iloc[0]).strip()
        if code.lower() in SKIP:
            continue
        # Kolom 6 = taxonomy-pad (bijv. "Bouwmaterialen - Gereedschap - ...")
        label = ""
        if len(row) > 6:
            raw = str(row.iloc[6]).strip()
            if raw not in ("nan", "#N/A", ""):
                label = raw
        if not label:
            label = code.replace("_", " ").title()
        result.append({"code": code, "label": label})
    return result


# Kolomgroepen in categorieindeling: (naam-kolom, id-kolom) paren (0-gebaseerd)
_CAT_GROEPEN = [(1, 3), (4, 6), (7, 9), (10, 12), (13, 15)]

# Rij-waarden die header zijn (overslaan)
_CAT_HEADERS = {
    "nl", "de", "type", "typ", "soort", "toepassing", "anwendung",
    "ondergrond", "substrat", "merken", "merke", "merk", "marke",
    "eigenschappen", "eigenschaft", "toepasing",
}


def laad_categorieindeling(pad: Path) -> list[dict]:
    """
    Laad categorieindeling uit het meerdere-sheets hiërarchische bestand.
    Elke sheet heeft kolomgroepen (naam, DE-naam, ID) op vaste posities.
    Geeft een platte lijst met id, naam en sectie.
    """
    if not pad.exists():
        return []

    skip_sheets = {"Legenda", "Acties"}
    result = []
    seen_ids: set[str] = set()

    xl = pd.ExcelFile(pad)
    for sheet_name in xl.sheet_names:
        if sheet_name in skip_sheets:
            continue
        df = pd.read_excel(pad, sheet_name=sheet_name, header=None)

        for _, row in df.iterrows():
            for naam_col, id_col in _CAT_GROEPEN:
                if naam_col >= len(row) or id_col >= len(row):
                    continue
                naam   = row.iloc[naam_col]
                id_val = row.iloc[id_col]

                if pd.isna(naam) or pd.isna(id_val):
                    continue
                naam   = str(naam).strip()
                id_str = str(id_val).strip()

                if naam.lower() in _CAT_HEADERS or not naam:
                    continue

                try:
                    id_int = int(float(id_str))
                    id_str = str(id_int)
                except (ValueError, OverflowError):
                    continue

                if id_str in seen_ids:
                    continue
                seen_ids.add(id_str)
                result.append({"id": id_str, "naam": naam, "sectie": sheet_name})

    return result


def _compacte_attribuutset_lijst(attribuutsets: list[dict]) -> str:
    return "\n".join(f"  {a['code']}: {a['label']}" for a in attribuutsets)


def _compacte_categorie_lijst(categorieen: list[dict]) -> str:
    by_sec: dict[str, list[dict]] = {}
    for c in categorieen:
        by_sec.setdefault(c.get("sectie", "Overig"), []).append(c)
    lines = []
    for sec, items in by_sec.items():
        lines.append(f"  [{sec}]")
        for c in items:
            lines.append(f"    {c['id']}: {c['naam']}")
    return "\n".join(lines)


# ─────────────────────────────────────────────
#  PIB / VIB LEZEN
# ─────────────────────────────────────────────

def _lees_pdf_tekst(pad: Path, max_tekens: int = 6000) -> str:
    """Extraheert leesbare tekst uit een PDF met pypdf."""
    try:
        reader = pypdf.PdfReader(str(pad))
        tekst_delen = []
        for pagina in reader.pages:
            tekst_delen.append(pagina.extract_text() or "")
            if sum(len(t) for t in tekst_delen) >= max_tekens:
                break
        return " ".join(" ".join(tekst_delen).split())[:max_tekens]
    except Exception:
        return ""


# Merken waarvan de PIB in een andere map staat (bijv. Sigma → PPG).
_PIB_MERK_ALIAS: dict[str, list[str]] = {
    "Sigma":       ["PPG"],
    "Sigmacover":  ["PPG"],
    "Sigmadur":    ["PPG"],
    "Sigmaweld":   ["PPG"],
}


def _zoek_pdf(basis_dir: Path, merk: str, omschrijving: str, prefix: str) -> Path | None:
    """
    Zoekt een PIB of VIB in {basis_dir}/{Merk}/NL/.
    prefix = "PIB" of "VIB"
    Probeert ook alias-mappen (bijv. Sigma → PPG).
    """
    merk_clean = _schone_bestandsnaam(merk)

    # Bouw kandidaat-mappen: eigen map + eventuele aliassen
    kandidaat_mappen: list[tuple[str, Path]] = []
    eigen_map = basis_dir / merk_clean / "NL"
    if eigen_map.exists():
        kandidaat_mappen.append((merk_clean, eigen_map))
    for alias in _PIB_MERK_ALIAS.get(merk, []):
        alias_clean = _schone_bestandsnaam(alias)
        alias_map   = basis_dir / alias_clean / "NL"
        if alias_map.exists():
            kandidaat_mappen.append((alias_clean, alias_map))

    if not kandidaat_mappen:
        return None

    # Exacte slug — strip brand alleen als gevolgd door spatie (niet bij "Sikaflex" → "Sika")
    omschr = omschrijving.strip()
    if merk and omschr.lower().startswith(merk.lower() + " "):
        omschr = omschr[len(merk):].strip()
    omschr_clean = _schone_bestandsnaam(omschr).strip(" -")

    for folder_merk, map_pad in kandidaat_mappen:
        exact = map_pad / f"{prefix} {folder_merk} {omschr_clean}-NL.pdf"
        if exact.exists():
            return exact

    # Fuzzy: brand + variant-suffixen strippen voor token-extractie.
    # - brand strip: "International One Up Web" → "One Up Web" (anders: tokens=["international"] → elke International PDF)
    # - variant strip: RAL-codes en basisaanduidingen staan nooit in PIB-bestandsnamen
    #   "Jotun Conseal TU RAL 8002 - B3" → "Conseal TU" → tokens=["conseal"]
    omschr_fuzzy = omschrijving.strip()
    if merk and omschr_fuzzy.lower().startswith(merk.lower() + " "):
        omschr_fuzzy = omschr_fuzzy[len(merk):].strip()
    # Strip kleurcode-systemen: RAL 7043 / RAL5024, JTN 1386, NCS S1234-... etc.
    omschr_fuzzy = re.sub(r"\s+(?:RAL|JTN|NCS)\s*[-/]?\s*\d+\S*", "", omschr_fuzzy, flags=re.IGNORECASE)
    # Strip basisaanduidingen: " - B3", "B1" aan het einde
    omschr_fuzzy = re.sub(r"\s*[-–]?\s*\bB[0-9]\b", "", omschr_fuzzy, flags=re.IGNORECASE)
    omschr_fuzzy = omschr_fuzzy.strip()

    def _bouw_tokens(tekst: str, min_tekst: int) -> list[str]:
        tokens = []
        for d in re.split(r"[\s\-_/®]+", tekst):
            d_lower = d.lower().strip("()[].")
            if not d_lower:
                continue
            if d_lower.isdigit() and len(d_lower) >= 2:
                tokens.append(d_lower)
            elif len(d_lower) >= min_tekst:
                tokens.append(d_lower)
        return tokens

    def _beste_match(kandidaten: list[Path]) -> Path | None:
        if not kandidaten:
            return None
        # Kortste naam = basisproduct; voorkomt dat een variant-PIB ("Flexi Alu") wint van de basis ("Flexi")
        return min(kandidaten, key=lambda p: len(p.name))

    # Pass 1: alle woorden ≥3 chars
    woorden = _bouw_tokens(omschr_fuzzy, 3)
    if woorden:
        for _, map_pad in kandidaat_mappen:
            matches = [k for k in map_pad.glob("*NL.pdf")
                       if all(w in k.name.lower() for w in woorden)]
            if matches:
                return _beste_match(matches)

    # Pass 2: alleen lange woorden (≥7 chars) + cijfers — filtert kleurcodesuffixen ("Yellow", "Green", "Set") weg.
    # "Sigmacover 280 Yellow/Green Set" → pass1: ["sigmacover","280","yellow","green","set"] → geen match
    # → pass2: ["sigmacover","280"] → matcht "PIB PPG Sigmacover 280-NL.pdf" ✓
    woorden_lang = _bouw_tokens(omschr_fuzzy, 7)
    if woorden_lang and woorden_lang != woorden:
        for _, map_pad in kandidaat_mappen:
            matches = [k for k in map_pad.glob("*NL.pdf")
                       if all(w in k.name.lower() for w in woorden_lang)]
            if matches:
                return _beste_match(matches)

    return None


def laad_zoekwoorden(pad: Path) -> list[dict]:
    if not pad or not pad.exists():
        return []
    import csv
    with open(pad, encoding="utf-8-sig") as f:
        return list(csv.DictReader(f, delimiter=";"))


_STOPWOORDEN = {"voor", "van", "met", "het", "een", "and", "the", "pro", "plus",
                "super", "ultra", "new", "all", "one", "max", "base", "top"}

def zoek_relevante_keywords(product: dict, zoekwoorden: list[dict], max_n: int = 6) -> list[str]:
    if not zoekwoorden:
        return []
    omschrijving = product.get("Omschrijving", "").lower()
    # Betekenisvolle woorden uit productnaam (≥4 chars, geen cijfers, geen stopwoorden)
    # Brand wordt NIET meegenomen — te generiek (brand matcht ook andere producttypes)
    woorden = [w for w in re.split(r"[\s\-/]+", omschrijving)
               if len(w) >= 4 and not w.isdigit() and w not in _STOPWOORDEN]

    resultaten = []
    for rij in zoekwoorden:
        doel = rij.get("doel_zoekwoord", "").lower()
        if not doel:
            continue
        doel_woorden = set(re.split(r"[\s\-/]+", doel))
        if any(w in doel_woorden for w in woorden):
            try:
                vol = int(rij.get("ads_zoekvolume") or 0)
            except ValueError:
                vol = 0
            resultaten.append((vol, doel))

    # Deduplicate, sorteer op volume
    gezien = set()
    uniek = []
    for vol, doel in sorted(resultaten, reverse=True):
        if doel not in gezien:
            gezien.add(doel)
            uniek.append(doel)
        if len(uniek) >= max_n:
            break
    return uniek


_WEBSHOP_DOMEINEN = {
    "bol.com", "beslist.nl", "amazon.", "coolblue.nl", "vidaxl.", "praxis.nl",
    "gamma.nl", "hornbach.nl", "karwei.nl", "marktplaats.nl", "google.", "youtube.",
    "facebook.", "instagram.", "twitter.", "linkedin.", "pinterest.", "fov.nl",
    "vergelijk.", "kiyoh.", "trustpilot.", "kieskeurig.", "tweakers.",
}


def _is_webshop(url: str) -> bool:
    url_lower = url.lower()
    return any(d in url_lower for d in _WEBSHOP_DOMEINEN)


def zoek_product_op_web(omschrijving: str, merk: str = "", max_urls: int = 3) -> list[dict]:
    """
    Zoekt op DuckDuckGo naar productinformatie en retourneert maximaal max_urls
    relevante (niet-webshop) pagina's als [{"url": ..., "titel": ..., "snippet": ...}].
    """
    if not _DDG_BESCHIKBAAR:
        return []

    query_delen = []
    if merk and merk.lower() not in omschrijving.lower():
        query_delen.append(merk)
    query_delen.append(omschrijving)
    query_delen.append("technische gegevens productinformatie")
    query = " ".join(query_delen)

    try:
        with _DDGS(verify=False) as ddg:
            resultaten = ddg.text(query, max_results=10)
    except Exception:
        return []

    gevonden = []
    for r in resultaten:
        url = r.get("href", "")
        if not url or _is_webshop(url):
            continue
        gevonden.append({
            "url":     url,
            "titel":   r.get("title", ""),
            "snippet": r.get("body", ""),
        })
        if len(gevonden) >= max_urls:
            break

    return gevonden


def _haal_pagina_tekst(url: str, max_tekens: int = 3000) -> str:
    """Haalt leesbare tekst op van een URL."""
    try:
        resp = httpx.get(url, timeout=12, follow_redirects=True,
                         verify=False, headers=_HEADERS)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        for tag in soup(["script", "style", "nav", "footer", "header"]):
            tag.decompose()
        return " ".join(soup.get_text(separator=" ").split())[:max_tekens]
    except Exception:
        return ""


# Domeinen die we voor afbeeldingen wél meenemen (concurrenten hebben goede productfoto's)
# maar voor tekst uitsluiten. Alleen social media / prijsvergelijkers / marktplaatsen buiten.
_AFBEELDING_UITSLUIT_DOMEINEN = {
    "marktplaats.nl", "amazon.", "google.", "youtube.",
    "facebook.", "instagram.", "twitter.", "linkedin.", "pinterest.",
    "vergelijk.", "kiyoh.", "trustpilot.", "kieskeurig.", "tweakers.",
    "beslist.nl", "fov.nl", "vidaxl.",
}


def _zoek_afbeelding_urls(omschrijving: str, merk: str = "", max_urls: int = 8) -> list[dict]:
    """
    Zoals zoek_product_op_web maar sluit webshops/concurrenten NIET uit —
    die hebben juist goede productfoto's op witte achtergrond.
    """
    if not _DDG_BESCHIKBAAR:
        return []
    query_delen = []
    if merk and merk.lower() not in omschrijving.lower():
        query_delen.append(merk)
    query_delen.append(omschrijving)
    query = " ".join(query_delen)
    try:
        with _DDGS(verify=False) as ddg:
            resultaten = ddg.text(query, max_results=15)
    except Exception:
        return []
    gevonden = []
    for r in resultaten:
        url = r.get("href", "")
        if not url:
            continue
        url_lower = url.lower()
        if any(d in url_lower for d in _AFBEELDING_UITSLUIT_DOMEINEN):
            continue
        gevonden.append({"url": url, "titel": r.get("title", ""), "snippet": r.get("body", "")})
        if len(gevonden) >= max_urls:
            break
    return gevonden


def haal_product_afbeelding(product: dict, client=None) -> tuple[list[str], str]:
    """
    Zoekt een productafbeelding via DuckDuckGo.
    Als client meegegeven is: haalt kandidaten op van meerdere pagina's en laat
    GPT-4o Vision de beste kiezen. Anders: eerste succesvolle pagina (origineel gedrag).
    Retourneert (paden, bron_log).
    """
    if not _DDG_BESCHIKBAAR:
        return [], ""

    omschrijving = product.get("Omschrijving", "")
    merk         = product.get("Merk", "")

    omschr_zoek = re.sub(r"\s+(?:RAL|JTN|NCS)\s*[-/]?\s*\d+\S*", "", omschrijving, flags=re.IGNORECASE)
    omschr_zoek = re.sub(r"\s*[-–]?\s*\bB[0-9]\b", "", omschr_zoek, flags=re.IGNORECASE).strip()

    urls = _zoek_afbeelding_urls(omschr_zoek, merk, max_urls=8)
    if not urls:
        return [], ""

    merk_clean = _schone_bestandsnaam(merk or "")
    omschr     = omschrijving.strip()
    if merk and omschr.lower().startswith(merk.lower() + " "):
        omschr = omschr[len(merk):].strip()
    # Strip RAL/JTN/B-codes en maatsuffix — varianten delen dezelfde blikfoto
    omschr = re.sub(r"\s+(?:RAL|JTN|NCS)\s*[-/]?\s*\d+\S*", "", omschr, flags=re.IGNORECASE)
    omschr = re.sub(r"\s*[-–]?\s*\bB[0-9]\b", "", omschr, flags=re.IGNORECASE)
    omschr = _strip_variant_suffix(omschr).strip()
    omschr_clean     = _schone_bestandsnaam(omschr).strip(" -")
    afbeelding_basis = f"{merk_clean.title()} {omschr_clean}" if omschr_clean else merk_clean.title()
    merk_dir         = AFBEELDING_BASE_DIR / merk_clean.lower().strip()

    if client:
        # Vision-modus: verzamel max 5 kandidaten van verschillende pagina's, AI kiest de beste
        zoekwoorden = [re.sub(r"[^a-z0-9]", "", merk_clean.lower())]
        for woord in omschr_zoek.lower().split():
            slug = re.sub(r"[^a-z0-9]", "", woord)
            if len(slug) > 2:
                zoekwoorden.append(slug)

        kandidaten: list[dict] = []
        for bron in urls[:6]:
            if len(kandidaten) >= 5:
                break
            resultaat = _kandidaat_bytes_van_pagina(bron["url"], zoekwoorden)
            if resultaat:
                img_bytes, img_url = resultaat
                kandidaten.append({"bytes": img_bytes, "page_url": bron["url"], "img_url": img_url})

        if not kandidaten:
            return [], ""

        beste_idx = _beoordeel_afbeeldingen_vision(kandidaten, client, omschrijving)
        beste     = kandidaten[beste_idx]
        save_pad  = merk_dir / f"{afbeelding_basis}.jpg"

        if _verwerk_afbeelding(beste["bytes"], save_pad):
            domein   = urlparse(beste["page_url"]).netloc
            bron_log = f"{len(kandidaten)} kandidaten beoordeeld, #{beste_idx + 1} gekozen ({domein})"
            return [str(save_pad)], bron_log
        return [], ""

    else:
        # Origineel gedrag: eerste succesvolle pagina
        for bron in urls:
            url = bron["url"]
            try:
                resp = httpx.get(url, timeout=15, follow_redirects=True,
                                 verify=False, headers=_HEADERS)
                resp.raise_for_status()
            except Exception:
                continue
            soup  = BeautifulSoup(resp.text, "html.parser")
            paden = _download_afbeeldingen(soup, url, merk_clean, afbeelding_basis)
            if paden:
                return paden, ""
        return [], ""


def haal_web_bronnen(product: dict) -> tuple[str, str]:
    """
    Zoekt productinformatie op het web als fallback wanneer geen PIB beschikbaar is.
    Retourneert (gecombineerde_tekst, log_string).
    """
    omschrijving = product.get("Omschrijving", "")
    merk         = product.get("Merk", "")

    urls = zoek_product_op_web(omschrijving, merk)
    if not urls:
        return "", ""

    tekst_delen = []
    gebruikte_bronnen = []

    for bron in urls:
        tekst = _haal_pagina_tekst(bron["url"])
        if len(tekst) < 200:
            # Pagina leeg of geblokkeerd — gebruik dan snippet van DDG
            tekst = bron["snippet"]
        if tekst:
            tekst_delen.append(f"[Bron: {bron['titel']} — {bron['url']}]\n{tekst}")
            gebruikte_bronnen.append(bron["url"])

    gecombineerd = "\n\n".join(tekst_delen)[:6000]
    log = f"Web: {len(gebruikte_bronnen)} pagina(s) — {', '.join(gebruikte_bronnen[:2])}"
    return gecombineerd, log


def lees_pib_vib_tekst(product: dict) -> tuple[str, str, str]:
    """
    Zoekt en leest de PIB en VIB voor een product.
    Geeft (pib_tekst, vib_tekst, gevonden_paden_log) terug.
    """
    merk         = str(product.get("Merk") or "")
    omschrijving = str(product.get("Omschrijving") or "")

    pib_pad = _zoek_pdf(config.PIB_BASE_DIR, merk, omschrijving, "PIB")
    vib_pad = _zoek_pdf(config.VIB_BASE_DIR, merk, omschrijving, "VIB")

    pib_tekst = _lees_pdf_tekst(pib_pad) if pib_pad else ""
    vib_tekst = _lees_pdf_tekst(vib_pad) if vib_pad else ""

    gevonden = []
    if pib_pad: gevonden.append(f"PIB: {pib_pad.name}")
    if vib_pad: gevonden.append(f"VIB: {vib_pad.name}")

    return pib_tekst, vib_tekst, " | ".join(gevonden)


# ─────────────────────────────────────────────
#  AI PROMPT
# ─────────────────────────────────────────────

def maak_prompt(product: dict, attribuutsets: list[dict], categorieen: list[dict], idx: int,
                pagina_tekst: str = "", pib_tekst: str = "", vib_tekst: str = "",
                zoekwoorden: list[str] | None = None) -> str:
    omschrijving = product.get("Omschrijving", "")
    merk         = product.get("Merk", "")
    artikelnr    = product.get("Artikelnummer", "")
    extra_info   = " | ".join(
        f"{k}: {v}" for k, v in product.items()
        if k not in ("Artikelnummer", "Omschrijving", "Merk",
                     "Productinformatieblad_pad", "Productveiligheidsblad_pad",
                     "URL_leverancier")
        and v and str(v) not in ("nan", "None")
    )
    if pagina_tekst:
        extra_info = (extra_info + " | " if extra_info else "") + f"Leverancierspagina: {pagina_tekst}"
    if pib_tekst:
        extra_info = (extra_info + " | " if extra_info else "") + f"Productinformatieblad: {pib_tekst}"
    if vib_tekst:
        extra_info = (extra_info + " | " if extra_info else "") + f"Veiligheidsinformatieblad: {vib_tekst[:1500]}"

    usp = config.SHOP_USPS[idx % len(config.SHOP_USPS)]  # niet meer in 1N1 gebruikt

    attrib_sectie = (
        _compacte_attribuutset_lijst(attribuutsets)
        if attribuutsets
        else "  (geen referentiebestand gevonden – gebruik beste schatting)"
    )
    cat_sectie = (
        _compacte_categorie_lijst(categorieen)
        if categorieen
        else "  (geen referentiebestand gevonden – gebruik beste schatting)"
    )

    if zoekwoorden:
        kw_instructie = (
            f"\n  - Gebruik bij voorkeur 1-2 van deze zoekwoorden uit onze database "
            f"als ze relevant zijn voor dit product: {', '.join(zoekwoorden)}"
        )
    else:
        kw_instructie = ""

    pib_instructie = ""
    if pib_tekst:
        pib_instructie = """
PIB-INSTRUCTIES (productinformatieblad beschikbaar):
- Haal de volledige verbruiks/opbrengsttabel over uit het PIB met ALLE ondergrondtypen en bijbehorende waarden. Render dit als HTML-tabel in het Verbruik-blok. Verzin dit NOOIT.
- Haal droogtijden LETTERLIJK over (bijv. "5-7 uur voor voetverkeer, 48-72 uur voor voertuigen").
- Haal toepassingsoppervlakken LETTERLIJK over (specifieke materialen zoals graniet, kalksteen, terracotta).
- 'Houdbaarheid' in het PIB = houdbaarheid van de verpakking (niet hoe lang de bescherming duurt). Noem dit NIET als productvoordeel.
- Belangrijke verwerkingswaarschuwingen (bijv. "niet verdunnen", "niet aanbrengen bij regen") verwerken in Verwerkingstips.
- Temperatuurbereik voor verwerking vermelden als dat in het PIB staat.
"""

    return f"""Je bent een productspecialist en SEO-expert voor {config.BEDRIJF_NAAM}. {config.BEDRIJF_NAAM} is {config.BEDRIJF_OMSCHRIJVING}.
{pib_instructie}
PRODUCT:
Artikelnummer: {artikelnr}
Omschrijving: {omschrijving}
Merk: {merk}
Extra: {extra_info or "geen"}

TAAK: Genereer alle KING-velden voor dit product als JSON.

ATTRIBUUTSET V4 – kies de BEST passende code uit deze lijst:
{attrib_sectie}

CATEGORIEINDELING – kies 1 tot 3 best passende IDs uit deze lijst:
{cat_sectie}

TAALCODE REGELS:
1N1 (meta titel):
  - Structuur: [MERK] [VOLLEDIGE PRODUCTNAAM] | FOV
  - Maximaal 2 segmenten gescheiden door één |, nooit meer
  - MINIMAAL 60, MAXIMAAL 70 tekens – tel exact
  - Als de productnaam + merk minder dan 57 tekens is: voeg de eenheid toe (bijv. "5 liter", "Can 10 liter") zodat je op 60-70 komt
  - Merk altijd vooraan, GEEN CTA-woorden (kopen/bestellen)
  - Eindigt exact op " | {config.BEDRIJF_NAAM}"

1N2 (meta description):
  - MINIMAAL 150, MAXIMAAL 160 tekens
  - Verwerk minimaal 2 concrete zoekwoorden die mensen gebruiken bij dit type product (bijv. materiaalnaam, toepassingsvorm, eigenschap){kw_instructie}
  - Geen herhaling van dezelfde gedachte in twee zinnen
  - Sluit af met een korte slotzin zoals "Bestel bij {config.BEDRIJF_NAAM}." of "Verkrijgbaar bij {config.BEDRIJF_NAAM}." — {config.BEDRIJF_NAAM} mag maar EENMAAL voorkomen in de gehele 1N2
  - GEEN "u" of "uw"

1NT (webtitel):
  - Volledige productnaam, leesbaar voor website
  - Zonder merk-prefix als merk al duidelijk is

1NF (factuuromschrijving):
  - Maximaal 50 tekens
  - Korte, duidelijke naam voor op de factuur

1NH (URL-slug):
  - Alles lowercase, woorden gescheiden door koppeltekens
  - Geen special characters, geen merk-prefix

1NL (lange omschrijving HTML):
  Gebruik EXACT deze structuur. Kopieer de style-tags letterlijk, vul de placeholders in.

  <style>
    .tab-style {{ display: block; padding: 1rem; border: 1px solid #ccc; background: white; margin-bottom: 1.5rem; border-radius: 5px; }} .tab-style h2 {{ margin-top: 0; color: #103a5d; }} .tab-style strong {{ color: #103a5d; }}
    .tabs {{ margin-top: 2rem; }} .tabs input[type="radio"] {{ display: none; }} .tabs label {{ padding: 0.5rem 1rem; background: #eee; margin-right: 0.2rem; cursor: pointer; border-top-left-radius: 5px; border-top-right-radius: 5px; font-weight: bold; color: #103a5d; }} .tabs label:hover {{ background: #ddd; }} .tabs .tab-content {{ display: none; border: 1px solid #ccc; padding: 1rem; background: white; border-top: none; border-radius: 0 0 5px 5px; }} .tabs input[type="radio"]:checked + label {{ background: #103a5d; color: white; }} .tabs input[type="radio"]:checked + label + .tab-content {{ display: block; }}
  </style>

  <h2>[Merk] [Productnaam] – [Korte ondertitel/toepassing]</h2>
  <p><strong>[Één zin: voor wie is dit product / hoofdtoepassing]</strong></p>
  <p>[Inleidende beschrijving in 2-3 zinnen over het product]</p>
  <ul>
    <li><span style="color:#c3a923;"><i class="check"></i></span> <strong>[Voordeel 1]</strong> – [toelichting]</li>
    <li><span style="color:#c3a923;"><i class="check"></i></span> <strong>[Voordeel 2]</strong> – [toelichting]</li>
    <li><span style="color:#c3a923;"><i class="check"></i></span> <strong>[Voordeel 3]</strong> – [toelichting]</li>
    <li><span style="color:#c3a923;"><i class="check"></i></span> <strong>[Voordeel 4]</strong> – [toelichting]</li>
    <li><span style="color:#c3a923;"><i class="check"></i></span> <strong>[Voordeel 5]</strong> – [toelichting]</li>
  </ul>
  <section style="margin-top: 3rem;">
    <h3>Waarom kiezen voor [productnaam]?</h3>
    <div style="display: flex; flex-wrap: wrap; gap: 1rem;">
      <div style="flex: 1; background-color: #f5f5f5; padding: 1rem;"><strong>[Kenmerk 1]</strong><p>[Toelichting]</p></div>
      <div style="flex: 1; background-color: #f5f5f5; padding: 1rem;"><strong>[Kenmerk 2]</strong><p>[Toelichting]</p></div>
      <div style="flex: 1; background-color: #f5f5f5; padding: 1rem;"><strong>[Kenmerk 3]</strong><p>[Toelichting]</p></div>
    </div>
  </section>
  <div style="display:flex;flex-wrap:wrap;gap:1rem;margin-top:2rem;">
    <div style="flex:1;background:#f5f5f5;padding:1rem;"><strong>Verbruik</strong>[Als het PIB een verbruikstabel per ondergrondtype heeft: render die VOLLEDIG als HTML-tabel: <table style="width:100%;border-collapse:collapse;font-size:13px;margin-top:0.5rem"><thead><tr><th style="text-align:left;border-bottom:1px solid #ccc;padding:4px 8px">Ondergrond</th><th style="text-align:left;border-bottom:1px solid #ccc;padding:4px 8px">m² per liter</th></tr></thead><tbody><tr><td style="padding:4px 8px">[ondergrond]</td><td style="padding:4px 8px">[waarde]</td></tr>...</tbody></table> — neem ALLE rijen over uit het PIB. Als er geen tabel is: één zin met het verbruik. Blok weglaten als verbruik volledig onbekend.]</div>
    <div style="flex:1;background:#f5f5f5;padding:1rem;"><strong>Droogtijd</strong><p>[droogtijd voor gebruik/voetverkeer/voertuigen — ALLEEN uit PIB, anders blok weglaten]</p></div>
    <div style="flex:1;background:#f5f5f5;padding:1rem;"><strong>Toepassing</strong><p>[verwerkingstemperatuur en -methode — ALLEEN uit PIB, anders blok weglaten]</p></div>
  </div>
  BELANGRIJK: laat een heel spec-blok weg als de waarde niet uit de productdata bekend is. Verzin NOOIT verbruik, droogtijd of technische specs.
  <div style="display:flex;flex-wrap:wrap;gap:1rem;margin-top:1rem;">
    <div style="flex:1;background:#f5f5f5;padding:1rem;">
      <strong>Waar toepassen?</strong>
      <ul><li>[toepassingsgebied 1]</li><li>[toepassingsgebied 2]</li><li>[toepassingsgebied 3]</li></ul>
    </div>
    <div style="flex:1;background:#f5f5f5;padding:1rem;">
      <strong>Gebruikstips</strong>
      <ul><li>[tip 1]</li><li>[tip 2]</li><li>[tip 3]</li></ul>
    </div>
  </div>
  <div class="tabs">
    <input type="radio" id="tab1_{artikelnr}" name="tabgroup_{artikelnr}" checked />
    <label for="tab1_{artikelnr}">Technische gegevens</label>
    <div class="tab-content"><p><strong>[spec label]:</strong> [waarde]</p><p><strong>[spec label]:</strong> [waarde]</p></div>
    <input type="radio" id="tab2_{artikelnr}" name="tabgroup_{artikelnr}" />
    <label for="tab2_{artikelnr}">Kenmerken</label>
    <div class="tab-content"><p>- [kenmerk 1]</p><p>- [kenmerk 2]</p><p>- [kenmerk 3]</p></div>
    <input type="radio" id="tab3_{artikelnr}" name="tabgroup_{artikelnr}" />
    <label for="tab3_{artikelnr}">Verwerkingstips</label>
    <div class="tab-content"><p>- [tip 1]</p><p>- [tip 2]</p><p>- [tip 3]</p></div>
  </div>

  Regels:
  - Minimaal 200 woorden totaal
  - Vul ALLE placeholders in met product-specifieke inhoud
  - Gebruik het artikelnummer {artikelnr} letterlijk in de tab id/name attributen
  - Schrijf in het Nederlands, geen "u"/"uw"
  - Geen verwijzing naar FOV aan het einde
  - NOOIT technische specs verzinnen (verbruik, droogtijd, treksterkte, temperatuur etc.) — alleen opnemen als ze expliciet in de productdata staan. Laat het blok of de waarde weg als de info ontbreekt.

GEEF ALLEEN geldige JSON (geen markdown, geen uitleg):
{{
  "attribuutset_code":       "...",
  "attribuutset_confidence": 0.85,
  "attribuutset_label":      "...",
  "categorie_ids":           "...,",
  "categorie_confidence":    0.80,
  "1N1": "...",
  "1N2": "...",
  "1NT": "...",
  "1NF": "...",
  "1NH": "...",
  "1NL": "..."
}}""".strip()


# ─────────────────────────────────────────────
#  OPENAI AANROEP
# ─────────────────────────────────────────────

def vraag_openai(prompt: str, client: OpenAI, model: str) -> dict:
    response = client.chat.completions.create(
        model=model,
        temperature=0.3,
        messages=[
            {
                "role": "system",
                "content": (
                    f"Je bent een {config.BEDRIJF_NAAM} productexpert en SEO-specialist. "
                    "Je schrijft altijd in het Nederlands. "
                    "Je telt tekens exact en houdt je aan alle lengteregels. "
                    "Antwoord ALLEEN in geldig JSON zonder markdown."
                )
            },
            {"role": "user", "content": prompt}
        ],
        max_tokens=4000,
        response_format={"type": "json_object"}
    )
    tekst = response.choices[0].message.content.strip()
    if tekst.startswith("```"):
        tekst = re.sub(r"^```json?\s*|\s*```$", "", tekst).strip()
    return json.loads(tekst)


def _fallback_waarden(product: dict) -> dict:
    omschrijving = product.get("Omschrijving", "")
    merk         = product.get("Merk", "FOV")
    slug = re.sub(r"[^a-z0-9]+", "-", f"{merk}-{omschrijving}".lower()).strip("-")
    return {
        "attribuutset_code":       "e_gereedschap_overig",
        "attribuutset_confidence": 0.10,
        "attribuutset_label":      "Handmatig invullen",
        "categorie_ids":           "",
        "categorie_confidence":    0.10,
        "1N1": f"{merk} {omschrijving} | {config.BEDRIJF_NAAM}"[:70],
        "1N2": f"{merk} {omschrijving}. Bestel bij {config.BEDRIJF_NAAM}.",
        "1NT": omschrijving,
        "1NF": f"{merk} {omschrijving}"[:50],
        "1NH": slug[:80],
        "1NL": f"<p>{omschrijving}</p>",
    }


def _conf_kleur(c: float) -> str:
    if c >= 0.75: return "D4EDDA"
    if c >= 0.55: return "FFF3CD"
    return "F8D7DA"


# ─────────────────────────────────────────────
#  VERWERK EEN PRODUCT
# ─────────────────────────────────────────────

def verwerk_product(
    product: dict,
    idx: int,
    attribuutsets: list[dict],
    categorieen: list[dict],
    client: OpenAI,
    model: str,
    log_func,
    zoekwoorden_data: list[dict] | None = None,
    merk_domeinen: dict | None = None,
    genereer_de: bool = False,
    genereer_jl: bool = False,
) -> dict:
    artikelnr = product.get("Artikelnummer", f"#{idx+1}")
    pagina    = {"tekst": "", "pib_url": "", "vib_url": ""}

    url = str(product.get("URL_leverancier", "") or "").strip()

    # Auto-detectie: zoek leveranciers-URL via merkdomein als geen URL opgegeven
    if not url.startswith("http") and merk_domeinen:
        gevonden = zoek_leverancier_url(
            str(product.get("Merk", "") or ""),
            str(product.get("Omschrijving", "") or ""),
            merk_domeinen,
        )
        if gevonden:
            url = gevonden
            log_func(f"  URL gevonden via merkdomein: {url[:70]}", "info")

    if url.startswith("http"):
        log_func(f"  Pagina ophalen: {url[:70]}", "info")
        pagina = haal_leverancier_pagina(
            url,
            merk=str(product.get("Merk") or ""),
            omschrijving=str(product.get("Omschrijving") or ""),
        )
        if pagina["pib_pad"]:
            log_func(f"  PIB opgeslagen: {pagina['pib_pad']}", "success")
        if pagina["vib_pad"]:
            log_func(f"  VIB opgeslagen: {pagina['vib_pad']}", "success")

    # PIB/VIB van K-schijf lezen
    pib_tekst, vib_tekst, pdf_log = lees_pib_vib_tekst(product)
    if pdf_log:
        log_func(f"  PDFs gelezen: {pdf_log}", "info")

    # Webzoek als fallback wanneer geen PIB beschikbaar is
    web_tekst = ""
    if not pib_tekst and not pagina["tekst"]:
        log_func(f"  Geen PIB gevonden — zoeken op web…", "info")
        web_tekst, web_log = haal_web_bronnen(product)
        if web_log:
            log_func(f"  {web_log}", "info")
        elif _DDG_BESCHIKBAAR:
            log_func(f"  Geen webresultaten gevonden", "warning")

    extra_tekst = pagina["tekst"] or web_tekst

    # Afbeelding zoeken via web als er nog geen beschikbaar is (URL_leverancier ontbreekt)
    if not pagina.get("afbeeldingen"):
        afb, afb_log = haal_product_afbeelding(product, client=client)
        if afb:
            log_msg = f"  Afbeelding: {Path(afb[0]).name}"
            if afb_log:
                log_msg += f" ({afb_log})"
            log_func(log_msg, "success")
        pagina["afbeeldingen"] = afb

    try:
        kw_matches = zoek_relevante_keywords(product, zoekwoorden_data or [])
        if kw_matches:
            log_func(f"  Zoekwoorden gevonden: {', '.join(kw_matches[:3])}{'…' if len(kw_matches) > 3 else ''}", "info")
        prompt = maak_prompt(product, attribuutsets, categorieen, idx, extra_tekst,
                             pib_tekst, vib_tekst, kw_matches or None)
        ai     = vraag_openai(prompt, client, model)

        # Vervang em/en-dashes door gewoon koppelteken (voorkomt encoding-problemen in KING)
        for k, v in ai.items():
            if isinstance(v, str):
                ai[k] = v.replace("—", " - ").replace("–", " - ")

        suffix = f" | {config.BEDRIJF_NAAM}"

        # Verwijder dubbele pipes (AI-artefact)
        ai["1N1"] = re.sub(r"\s*\|\s*\|\s*", " | ", ai.get("1N1", ""))

        # Zorg dat 1N1 eindigt op " | {BEDRIJF_NAAM}" en max 2 segmenten heeft
        n1 = ai.get("1N1", "")
        segmenten = [s.strip() for s in n1.split("|")]
        n1_body = segmenten[0].strip()

        # Verwijder dubbel merk: "Hempel Hempel's..." → "Hempel's..."
        merk = str(product.get("Merk", "") or "").strip()
        if merk:
            merk_lower = merk.lower()
            body_lower = n1_body.lower()
            if body_lower.startswith(merk_lower + " " + merk_lower):
                n1_body = n1_body[len(merk) + 1:].strip()

        # Eigenmerkartikel: merk == bedrijfsnaam → "FOV Plamuurmes | FOV" is redundant
        if merk and merk.lower() == config.BEDRIJF_NAAM.lower():
            if n1_body.lower().startswith(merk.lower() + " "):
                n1_body = n1_body[len(merk):].strip()

        n1 = n1_body + suffix  # altijd 2 segmenten

        # Te kort: probeer eenheid toe te voegen
        if len(n1) < 60:
            eenheid = str(product.get("Eenheid", "") or "").strip()
            eenheid_woorden = [w for w in eenheid.lower().split() if len(w) >= 3]
            al_aanwezig = any(w in n1.lower() for w in eenheid_woorden)
            if eenheid and not al_aanwezig:
                n1_met = n1_body + " " + eenheid + suffix  # n1_body, niet segmenten[0]
                if len(n1_met) <= 70:
                    n1 = n1_met

        # Te lang: afkappen voor suffix
        if len(n1) > 70:
            n1 = n1[:70 - len(suffix)].rsplit(" ", 1)[0] + suffix

        ai["1N1"] = n1

        # Verwijder dubbele FOV-vermelding in 1N2
        n2 = ai.get("1N2", "")
        bedrijf_lower = config.BEDRIJF_NAAM.lower()
        n2_zinnen = re.split(r"(?<=[.!?])\s+", n2.strip())
        # Bewaar alleen de eerste zin die BEDRIJF_NAAM bevat; verwijder duplicaten
        fov_gezien = False
        gefilterd = []
        for zin in n2_zinnen:
            if bedrijf_lower in zin.lower():
                if not fov_gezien:
                    gefilterd.append(zin)
                    fov_gezien = True
            else:
                gefilterd.append(zin)
        n2 = " ".join(gefilterd)

        # Lengte 1N2 controleren — afkappen op laatste volledige zin binnen 160 tekens
        if len(n2) > 160:
            zinnen = re.split(r"(?<=[.!?])\s+", n2.strip())
            passend = ""
            for zin in zinnen:
                kandidaat = (passend + " " + zin).strip() if passend else zin
                if len(kandidaat) <= 160:
                    passend = kandidaat
                else:
                    break
            n2 = passend if passend else n2[:160].rsplit(" ", 1)[0]

        # Losse check: te kort → aanvullen (ook na truncatie), maar nooit boven 160
        if len(n2) < 150 and n2:
            if bedrijf_lower not in n2.lower():
                kandidaat = n2.rstrip(".") + f". Bestel snel bij {config.BEDRIJF_NAAM}."
            else:
                stripped = n2.rstrip()
                sep = " " if stripped.endswith(".") else ". "
                kandidaat = stripped + sep + "Snel geleverd."
            if len(kandidaat) <= 160:
                n2 = kandidaat

        ai["1N2"] = n2

        log_func(f"✓ {artikelnr} – 1N1: {len(ai.get('1N1',''))}t, 1N2: {len(ai.get('1N2',''))}t", "success")

        if genereer_de:
            log_func(f"  DE taalcodes genereren…", "info")
            de = genereer_de_taalcodes(product, ai, client)
            ai.update(de)
            log_func(f"  DE klaar – 4D1: {len(de.get('4D1_FOV_DE_META_DATA_1',''))}t", "success")

        if genereer_jl:
            log_func(f"  Jachtlakken taalcodes genereren…", "info")
            jl = genereer_jl_taalcodes(product, ai, client)
            ai.update(jl)
            log_func(f"  JL klaar – 2N1: {len(jl.get('2N1_JL_NL_META_DATA_1',''))}t", "success")

        return {**product, "_ai": ai, "_pagina": pagina, "_status": "success"}

    except Exception as e:
        log_func(f"✗ {artikelnr} – {str(e)[:80]}", "error")
        return {**product, "_ai": _fallback_waarden(product), "_pagina": pagina, "_status": "error"}


# ─────────────────────────────────────────────
#  EXCEL OPBOUWEN
# ─────────────────────────────────────────────

def bouw_excel(
    producten: list[dict],
    output_pad: Path,
    client: OpenAI,
    model: str,
    attribuutsets: list[dict],
    categorieen: list[dict],
    log_func,
    progress_func,
    zoekwoorden_data: list[dict] | None = None,
    merk_domeinen: dict | None = None,
    output_json_pad: Path | None = None,
    genereer_de: bool = False,
    genereer_jl: bool = False,
) -> Path:
    kolommen = KOLOMMEN + (DE_KOLOMMEN if genereer_de else []) + (JL_KOLOMMEN if genereer_jl else [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KING Import"

    hdr_font  = Font(name="Arial", bold=True, color=config.CLR_WIT, size=10)
    hdr_fill  = PatternFill("solid", fgColor=config.CLR_BLAUW)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_font  = Font(name="Arial", size=9)
    data_align = Alignment(vertical="top", wrap_text=True)
    rand = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for ci, kol in enumerate(kolommen, 1):
        c = ws.cell(row=1, column=ci, value=kol)
        c.font = hdr_font; c.fill = hdr_fill
        c.alignment = hdr_align; c.border = rand
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    verwerkte_producten = []
    _alle_waarden: list[dict] = []

    for i, product in enumerate(producten):
        progress_func(i + 1, len(producten))
        result = verwerk_product(product, i, attribuutsets, categorieen, client, model, log_func,
                                 zoekwoorden_data=zoekwoorden_data,
                                 merk_domeinen=merk_domeinen or {},
                                 genereer_de=genereer_de,
                                 genereer_jl=genereer_jl)
        verwerkte_producten.append(result)

        if i < len(producten) - 1:
            time.sleep(0.5)

    # Kopieer 1NL van parent naar children (variantengroep)
    parent_ai = {
        r["Artikelnummer"]: r.get("_ai", {})
        for r in verwerkte_producten
        if r.get("_type") == "parent"
    }
    for result in verwerkte_producten:
        if result.get("_type") == "child":
            parent_code = result.get("_parent_code", "")
            if parent_code in parent_ai:
                result["_ai"]["1NL"] = parent_ai[parent_code].get("1NL", result["_ai"].get("1NL", ""))

    for ri, result in enumerate(verwerkte_producten, 2):
        ai  = result.get("_ai", _fallback_waarden(result))
        conf = float(ai.get("attribuutset_confidence", 0.10))
        cf   = PatternFill("solid", fgColor=_conf_kleur(conf))

        pagina  = result.get("_pagina", {})

        # Fix 1: placeholder-waarden uit template negeren
        def _echte_pad(waarde):
            v = str(waarde or "").strip()
            return "" if v.startswith(r"C:\paden") or "leverancierspagina" in v else v

        def _lokaal_pad(waarde, prefix: str) -> str:
            """Als waarde een URL is, download dan naar lokale mapstructuur."""
            v = _echte_pad(waarde)
            if not v or not v.startswith("http"):
                return v
            merk_c = _schone_bestandsnaam(str(result.get("Merk") or ""))
            omschr = str(result.get("Omschrijving") or "")
            if omschr.lower().startswith(merk_c.lower()):
                omschr = omschr[len(merk_c):].strip()
            omschr_c = _schone_bestandsnaam(omschr).strip(" -")
            merk_dir = config.PIB_BASE_DIR / merk_c / "NL"
            bestand = merk_dir / f"{prefix} {merk_c} {omschr_c}-NL.pdf"
            return str(bestand) if _download_pdf(v, bestand) else v

        pib_pad = _lokaal_pad(result.get("Productinformatieblad_pad"), "PIB") or pagina.get("pib_pad", "")
        vib_pad = _lokaal_pad(result.get("Productveiligheidsblad_pad"), "VIB") or pagina.get("vib_pad", "")
        afbeeldingen = pagina.get("afbeeldingen", [])

        waarden = {
            "Artikelnummer":                        result.get("Artikelnummer", ""),
            "Eenheid":                              result.get("Eenheid", "Stuk"),
            "Zoekcode":                             result.get("Merk", ""),
            "Omschrijving":                         result.get("Omschrijving", ""),
            "Opbrengstgroep":                       result.get("Opbrengstgroep", ""),
            "WebArtikel":                           config.WEBART,
            "TekstOpFactuur":                       result.get("Omschrijving", ""),
            "AfbeeldingKlein":                      afbeeldingen[0] if afbeeldingen else "",
            "AfbeeldingGroot":                      afbeeldingen[0] if afbeeldingen else "",
            "Leveranciernummer":                    result.get("Leveranciernummer", ""),
            "Leveranciernaam":                      result.get("Merk", ""),
            "ArtikelOmschrijvingLeverancier":       result.get("Omschrijving", ""),
            "ArtikelNummerBijLeverancier":          result.get("ArtikelNummerBijLeverancier", "") or "",
            "EanCode":                              result.get("EanCode", "") or "",
            "VR_ART_Magentotype":                   "Configureerbaar" if result.get("_type") == "parent" else "Simpel",
            "VR_ART_Zichtbaarheid":                 "Niet individueel zichtbaar" if result.get("_type") == "child" else "Catalogus, zoeken",
            "VR_ART_Virtueel_Artikel_SKU":          result.get("_parent_code", "") if result.get("_type") == "child" else (result.get("Artikelnummer", "") if result.get("_type") == "parent" else ""),
            "VR_ART_Configuratie_op":              result.get("_configuratie_op", ""),
            "VR_ART_Actief_in_shop":                config.ACTIEF_IN_SHOP,
            "VR_ART_F-Merk":                        result.get("Merk", ""),
            "VR_ART_Extra_afbeelding_1":            afbeeldingen[1] if len(afbeeldingen) > 1 else "",
            "VR_ART_Extra_afbeelding_2":            afbeeldingen[2] if len(afbeeldingen) > 2 else "",
            "VR_ART_Extra_afbeelding_3":            afbeeldingen[3] if len(afbeeldingen) > 3 else "",
            "VR_ART_Extra_afbeelding_4":            afbeeldingen[4] if len(afbeeldingen) > 4 else "",
            "VR_ART_Extra_afbeelding_5":            afbeeldingen[5] if len(afbeeldingen) > 5 else "",
            "VR_ART_Productinformatieblad_NL":      pib_pad,
            "VR_ART_Productveiligheidsblad_NL_A":   vib_pad,
            "VR_ART_Attribuutset_V4":               ai.get("attribuutset_code", ""),
            "1NT_FOV_NL_TITLE_WEB":                 ai.get("1NT", ""),
            "1NF_FOV_NL_TITLE_FACTUUR":             ai.get("1NF", ""),
            "1NH_FOV_NL_URL":                       ai.get("1NH", ""),
            "1N1_FOV_NL_META_DATA_1":               ai.get("1N1", ""),
            "1N2_FOV_NL_META_DATA_2":               ai.get("1N2", ""),
            "1NL_FOV_NL_LANGE_OMSCHRIJVING":        ai.get("1NL", ""),
            "Attribuutset_Confidence_%":            f"{conf*100:.0f}%",
            "Attribuutset_Label":                   ai.get("attribuutset_label", ""),
            "VR_ART_Webcategorie_ID._V2":           ai.get("categorie_ids", ""),
            "Webcategorie_Confidence_%":            f"{float(ai.get('categorie_confidence', 0))*100:.0f}%",
            # DE taalcodes (alleen gevuld als genereer_de=True)
            "4D1_FOV_DE_META_DATA_1":              ai.get("4D1_FOV_DE_META_DATA_1", ""),
            "4D2_FOV_DE_META_DATA_2":              ai.get("4D2_FOV_DE_META_DATA_2", ""),
            "4DF_FOV_DE_TITLE_FACTUUR":            ai.get("4DF_FOV_DE_TITLE_FACTUUR", ""),
            "4DT_FOV_DE_TITLE_WEB":                ai.get("4DT_FOV_DE_TITLE_WEB", ""),
            "4DU_FOV_DE_URL":                      ai.get("4DU_FOV_DE_URL", ""),
            "4DL_FOV_DE_LANGE_OMSCHRIJVING":       ai.get("4DL_FOV_DE_LANGE_OMSCHRIJVING", ""),
            # JL taalcodes (alleen gevuld als genereer_jl=True)
            "2N1_JL_NL_META_DATA_1":               ai.get("2N1_JL_NL_META_DATA_1", ""),
            "2N2_JL_NL_META_DATA_2":               ai.get("2N2_JL_NL_META_DATA_2", ""),
            "2NF_JL_NL_TITLE_FACTUUR":             ai.get("2NF_JL_NL_TITLE_FACTUUR", ""),
            "2NT_JL_NL_TITLE_WEB":                 ai.get("2NT_JL_NL_TITLE_WEB", ""),
            "2NH_JL_NL_URL":                       ai.get("2NH_JL_NL_URL", ""),
            "2NL_JL_NL_LANGE_OMSCHRIJVING":        ai.get("2NL_JL_NL_LANGE_OMSCHRIJVING", ""),
        }

        for ci, kol in enumerate(kolommen, 1):
            cel = ws.cell(row=ri, column=ci, value=waarden.get(kol, ""))
            cel.font = data_font; cel.alignment = data_align; cel.border = rand
            if kol in HANDMATIG:
                cel.fill = PatternFill("solid", fgColor=config.CLR_ORANJE)
            elif kol in JL_VELDEN:
                cel.fill = PatternFill("solid", fgColor="C9EFF1")
            elif kol in AI_VELDEN or kol in DE_VELDEN:
                cel.fill = PatternFill("solid", fgColor=config.CLR_GROEN)
            elif kol in MAPPING_VELDEN:
                cel.fill = cf

        _alle_waarden.append(waarden)

    # Kolombreedte
    breedte = {
        "1N1_FOV_NL_META_DATA_1": 44, "1N2_FOV_NL_META_DATA_2": 58,
        "1NF_FOV_NL_TITLE_FACTUUR": 34, "1NH_FOV_NL_URL": 38,
        "1NL_FOV_NL_LANGE_OMSCHRIJVING": 88, "1NT_FOV_NL_TITLE_WEB": 28,
        "Omschrijving": 34, "Artikelnummer": 18,
        "VR_ART_Productinformatieblad_NL": 64, "VR_ART_Productveiligheidsblad_NL_A": 64,
        "VR_ART_Attribuutset_V4": 30, "Attribuutset_Confidence_%": 14,
        "Attribuutset_Label": 26, "VR_ART_Webcategorie_ID._V2": 24,
        "Webcategorie_Confidence_%": 14,
    }
    for ci, kol in enumerate(kolommen, 1):
        ws.column_dimensions[get_column_letter(ci)].width = breedte.get(kol, 17)

    _voeg_legenda_toe(wb, attribuutsets)
    wb.save(output_pad)

    if output_json_pad and _alle_waarden:
        with open(output_json_pad, "w", encoding="utf-8") as _f:
            json.dump(_alle_waarden, _f, ensure_ascii=False, default=str)

    log_func(f"Klaar — {len(producten)} producten → {output_pad.name}", "success")
    return output_pad


def _voeg_legenda_toe(wb: openpyxl.Workbook, attribuutsets: list[dict]):
    ws2 = wb.create_sheet("Legenda")
    ws2.merge_cells("A1:B1")
    ws2["A1"].value = "Legenda – Product Import"
    ws2["A1"].font  = Font(name="Arial", bold=True, size=13, color=config.CLR_WIT)
    ws2["A1"].fill  = PatternFill("solid", fgColor=config.CLR_BLAUW)
    ws2.row_dimensions[1].height = 24

    items = [
        ("KLEURCODERING", None),
        ("Oranje", "Handmatig invullen (prijs, leveranciernummer)"),
        ("Lichtgroen", "AI-gegenereerde teksten – altijd controleren vóór publicatie"),
        ("Groen (mapping)", "Hoge confidence ≥ 75% – waarschijnlijk correct"),
        ("Geel (mapping)", "Matige confidence 55-74% – controleer de waarde"),
        ("Rood (mapping)", "Lage confidence < 55% – handmatig aanpassen"),
        ("", None),
        ("INPUTKOLOMMEN (verplicht)", None),
        ("Artikelnummer", "KING artikelnummer"),
        ("Omschrijving", "Productnaam / omschrijving"),
        ("Merk", "F-merk (bijv. Jotun, Hempel, Sika)"),
        ("", None),
        ("INPUTKOLOMMEN (optioneel)", None),
        ("EanCode", "EAN/barcode"),
        ("Eenheid", "Stuk, Liter, etc."),
        ("Gewicht", "In kg"),
        ("VerkoopPrijsExBTW / AdviesPrijsExBTW", "Prijzen (oranje = handmatig invullen)"),
        ("Leveranciernummer", "KING leveranciernummer"),
        ("ArtikelNummerBijLeverancier", "Bestelnummer bij leverancier"),
        ("Productinformatieblad_pad", r"K:\pad\naar\PIB.pdf"),
        ("Productveiligheidsblad_pad", r"K:\pad\naar\VIB.pdf"),
        ("", None),
        ("AI-VELDEN (lichtgroen)", None),
        ("1N1", "Meta titel: Merk Product | USP | [BEDRIJF] (60-70 tekens)"),
        ("1N2", "Meta description (150-160 tekens)"),
        ("1NT", "Webtitel (zichtbare naam op website)"),
        ("1NF", "Factuurnaam (max 50 tekens)"),
        ("1NH", "URL-slug (alles na fov.nl/)"),
        ("1NL", "Lange HTML omschrijving"),
        ("", None),
        ("AUTO-MAPPING (kleurgecodeerd)", None),
        ("VR_ART_Attribuutset_V4", "Uit attribuutset_v4.xlsx – zie referentiedata map"),
        ("VR_ART_Webcategorie_ID._V2", "Uit categorieindeling.xlsx – kommagescheiden IDs"),
        ("Confidence_%", "< 55% = rood = handmatig controleren"),
        ("", None),
        ("VASTE WAARDEN (automatisch)", None),
        ("WebArtikel", "1 (live in shop)"),
        ("VR_ART_Actief_in_shop", "1 (online)"),
        ("VR_ART_F-Merk", "Overgenomen uit Merk kolom"),
    ]

    for i, (k, v) in enumerate(items, 3):
        ck = ws2.cell(row=i, column=1, value=k)
        ck.font = Font(name="Arial", size=9,
                       bold=(v is None and bool(k)),
                       color=config.CLR_BLAUW if v is None and k else "000000")
        if v is None and k:
            ck.fill = PatternFill("solid", fgColor="E8EFF7")
        if v:
            ws2.cell(row=i, column=2, value=v).font = Font(name="Arial", size=9)

    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 72

    # Attribuutset referentie op apart tabblad
    if attribuutsets:
        ws3 = wb.create_sheet("Attribuutsets")
        ws3.cell(1, 1, "Code").font = Font(bold=True, name="Arial")
        ws3.cell(1, 2, "Label").font = Font(bold=True, name="Arial")
        for i, a in enumerate(attribuutsets, 2):
            ws3.cell(i, 1, a["code"])
            ws3.cell(i, 2, a["label"])
        ws3.column_dimensions["A"].width = 35
        ws3.column_dimensions["B"].width = 40


# ─────────────────────────────────────────────
#  INVOER INLEZEN
# ─────────────────────────────────────────────

_KOLOM_ALIASSEN = {
    "Artikelomschrijving": "Omschrijving",
    "Artikelnaam":         "Omschrijving",
    "ProductNaam":         "Omschrijving",
    "EAN":                 "EanCode",
    "EAN-code":            "EanCode",
    "Leverancier":         "Merk",
    "F-merk":              "Merk",
}


def lees_invoer(pad: Path) -> list[dict]:
    """Lees Excel of CSV invoerbestand naar lijst van dicts."""
    suffix = pad.suffix.lower()
    if suffix == ".csv":
        # Auto-detect separator: als meer puntkomma's dan komma's in regel 1 → puntkomma
        with open(pad, encoding="utf-8-sig", errors="replace") as fh:
            eerste_regel = fh.readline()
        sep = ";" if eerste_regel.count(";") > eerste_regel.count(",") else ","
        df = pd.read_csv(pad, encoding="utf-8-sig", sep=sep)
    elif suffix in (".xlsx", ".xls"):
        df = pd.read_excel(pad)
    else:
        raise ValueError(f"Niet-ondersteund bestandsformaat: {suffix}")

    df.rename(columns=_KOLOM_ALIASSEN, inplace=True)

    import math
    df.columns = [str(c).strip() for c in df.columns]
    df = df.dropna(subset=["Artikelnummer"])
    records = df.to_dict("records")
    return [
        {k: (None if (v is None or (isinstance(v, float) and math.isnan(v))) else v)
         for k, v in row.items()}
        for row in records
    ]


# ─────────────────────────────────────────────
#  STANDALONE GEBRUIK
# ─────────────────────────────────────────────

if __name__ == "__main__":
    from openai import OpenAI

    if not config.OPENAI_API_KEY:
        print("Stel OPENAI_API_KEY in als omgevingsvariabele.")
        raise SystemExit(1)

    bron_bestanden = list(config.BRON_DIR.glob("*.xlsx")) + list(config.BRON_DIR.glob("*.csv"))
    if not bron_bestanden:
        print(f"Geen invoerbestanden gevonden in {config.BRON_DIR}")
        raise SystemExit(1)

    invoer_pad = bron_bestanden[0]
    print(f"Invoer: {invoer_pad.name}")

    producten    = lees_invoer(invoer_pad)
    attribuutsets = laad_attribuutset(config.ATTRIBUUTSET_FILE)
    categorieen   = laad_categorieindeling(config.CATEGORIE_FILE)
    client       = OpenAI(api_key=config.OPENAI_API_KEY,
                          http_client=httpx.Client(verify=False))

    print(f"Producten: {len(producten)}")
    print(f"Attribuutsets geladen: {len(attribuutsets)}")
    print(f"Categorieen geladen:   {len(categorieen)}")

    datum      = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_pad = config.OUTPUT_DIR / f"KING_import_{datum}.xlsx"

    totaal = [0]

    import sys
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    def log(msg, t="info"):
        print(msg)

    def progress(h, t):
        if h != totaal[0]:
            totaal[0] = h
            print(f"  [{h}/{t}]")

    bouw_excel(producten, output_pad, client, config.OPENAI_MODEL,
               attribuutsets, categorieen, log, progress)
    print(f"\nOutput: {output_pad}")

