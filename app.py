"""
Product Import Generator – Flask Web App
==========================================
Start: python app.py
Open:  http://127.0.0.1:5002
"""

import io
import json
import os
import ssl
import threading
import time
from datetime import datetime
from pathlib import Path

import httpx
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from flask import Flask, jsonify, render_template, request, send_file
from openai import OpenAI

import config
from nieuw_product_generator import (
    bouw_excel,
    detecteer_variantgroepen,
    laad_attribuutset,
    laad_categorieindeling,
    laad_merk_domeinen,
    laad_zoekwoorden,
    lees_invoer,
)
from king_artikel_push import KingConfig, PushDB, push_batch, huidig_gebruiker

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "change-me-in-production")

# ─────────────────────────────────────────────
#  GLOBALE STATE
# ─────────────────────────────────────────────

progress_data = {
    "running": False, "current": 0, "total": 0,
    "log": [], "error": None, "done": False,
    "output_bestand": None,
    "waarden_bestand": None,
}
progress_lock = threading.Lock()

push_data: dict = {
    "running": False, "current": 0, "total": 0,
    "log": [], "error": None, "done": False,
    "sessie_id": None, "n_gelukt": 0, "n_mislukt": 0,
}
push_lock = threading.Lock()

_push_db: PushDB | None = None


def _get_push_db() -> PushDB:
    global _push_db
    if _push_db is None:
        _push_db = PushDB(config.PUSH_DB_FILE)
    return _push_db


def _king_config_from_env() -> KingConfig | None:
    if not (config.KING_HOST and config.KING_ACCESS_TOKEN and config.KING_ADMINISTRATIE):
        return None
    return KingConfig(
        protocol=config.KING_PROTOCOL,
        host=config.KING_HOST,
        port=config.KING_PORT,
        administratie=config.KING_ADMINISTRATIE,
        access_token=config.KING_ACCESS_TOKEN,
        sql_timeout=config.KING_SQL_TIMEOUT,
        verify_ssl=False,
    )


def log(msg: str, type: str = "info"):
    with progress_lock:
        progress_data["log"].append({
            "msg": msg, "type": type,
            "time": datetime.now().strftime("%H:%M:%S"),
        })


def reset_progress():
    with progress_lock:
        progress_data.update({
            "running": True, "current": 0, "total": 0,
            "log": [], "error": None, "done": False,
            "output_bestand": None,
        })


def progress_func(huidig: int, totaal: int):
    with progress_lock:
        progress_data["current"] = huidig
        progress_data["total"]   = totaal


# ─────────────────────────────────────────────
#  ROUTES
# ─────────────────────────────────────────────

@app.route("/")
def index():
    attrib_ok = config.ATTRIBUUTSET_FILE.exists()
    cat_ok    = config.CATEGORIE_FILE.exists()
    return render_template("index.html",
                           attrib_ok=attrib_ok,
                           cat_ok=cat_ok,
                           attrib_pad=str(config.ATTRIBUUTSET_FILE),
                           cat_pad=str(config.CATEGORIE_FILE))


@app.route("/upload", methods=["POST"])
def upload():
    try:
        f = request.files.get("invoer")
        if not f:
            return jsonify({"ok": False, "error": "Geen bestand meegestuurd"})

        ext = Path(f.filename).suffix.lower()
        if ext not in (".xlsx", ".xls", ".csv"):
            return jsonify({"ok": False, "error": "Alleen .xlsx, .xls of .csv"})

        pad = config.UPLOAD_DIR / f"invoer{ext}"
        f.save(pad)

        producten = lees_invoer(pad)
        if not producten:
            return jsonify({"ok": False, "error": "Geen producten gevonden (controleer Artikelnummer kolom)"})

        # Controleer verplichte kolommen
        missende = [k for k in ("Artikelnummer", "Omschrijving", "Merk") if k not in producten[0]]
        if missende:
            return jsonify({"ok": False, "error": f"Missende kolommen: {', '.join(missende)}"})

        # Auto-detectie variant-groepen
        producten = detecteer_variantgroepen(producten)

        # Samenvatting voor preview
        n_parent   = sum(1 for p in producten if p.get("_type") == "parent")
        n_child    = sum(1 for p in producten if p.get("_type") == "child")
        n_los      = sum(1 for p in producten if not p.get("_type"))
        groep_ids  = sorted({p["_groep_id"] for p in producten if p.get("_groep_id")})

        groepen_preview = []
        for gid in groep_ids:
            leden = [p for p in producten if p.get("_groep_id") == gid]
            parent = next((p for p in leden if p.get("_type") == "parent"), None)
            children = [p for p in leden if p.get("_type") == "child"]
            groepen_preview.append({
                "parent_nr":  parent["Artikelnummer"] if parent else "?",
                "parent_oms": parent["Omschrijving"]  if parent else "?",
                "children":   [{"nr": c["Artikelnummer"], "oms": c["Omschrijving"]} for c in children],
            })

        preview = [
            {
                "Artikelnummer": p.get("Artikelnummer", ""),
                "Omschrijving":  p.get("Omschrijving", ""),
                "Merk":          p.get("Merk", ""),
                "Type":          p.get("_type", "los").capitalize(),
            }
            for p in producten[:12]
        ]

        with open(config.UPLOAD_DIR / "werkdocument.json", "w", encoding="utf-8") as fp:
            json.dump(producten, fp, ensure_ascii=False, default=str)

        return jsonify({
            "ok":      True,
            "count":   len(producten),
            "n_parent": n_parent,
            "n_child":  n_child,
            "n_los":    n_los,
            "groepen":  groepen_preview,
            "preview":  preview,
        })
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/start", methods=["POST"])
def start():
    wp = config.UPLOAD_DIR / "werkdocument.json"
    if not wp.exists():
        return jsonify({"ok": False, "error": "Upload eerst een invoerbestand (stap 1)"})

    if progress_data["running"] and not progress_data["done"]:
        return jsonify({"ok": False, "error": "Verwerking al bezig – wacht of klik Stop"})

    model        = request.json.get("model", config.OPENAI_MODEL) if request.json else config.OPENAI_MODEL
    genereer_de  = True
    genereer_jl  = bool((request.json or {}).get("genereer_jl", False))
    api_key = config.OPENAI_API_KEY
    if not api_key:
        return jsonify({"ok": False, "error": "OPENAI_API_KEY omgevingsvariabele niet ingesteld"})

    try:
        with open(wp, encoding="utf-8") as f:
            producten = json.load(f)
    except Exception as e:
        return jsonify({"ok": False, "error": f"Werkdocument lezen mislukt: {e}"})

    reset_progress()
    extra = " | +DE" + (" | +Jachtlakken" if genereer_jl else "")
    log(f"Start – {len(producten)} producten | model: {model}{extra}", "info")

    attribuutsets    = laad_attribuutset(config.ATTRIBUUTSET_FILE)
    categorieen      = laad_categorieindeling(config.CATEGORIE_FILE)
    zoekwoorden_data = laad_zoekwoorden(config.ZOEKWOORDEN_FILE) if config.ZOEKWOORDEN_FILE else []
    merk_domeinen    = laad_merk_domeinen(config.MERK_DOMEINEN_FILE)

    if attribuutsets:
        log(f"Attribuutset V4 geladen: {len(attribuutsets)} codes", "info")
    else:
        log(f"Attribuutset referentie niet gevonden – AI maakt eigen schatting", "warning")

    if categorieen:
        log(f"Categorieindeling geladen: {len(categorieen)} categorieën", "info")
    else:
        log(f"Categorieindeling niet gevonden – AI maakt eigen schatting", "warning")

    if merk_domeinen:
        log(f"Merkdomeinen geladen: {len(merk_domeinen)} merken", "info")

    client = OpenAI(
        api_key=api_key,
        http_client=httpx.Client(verify=False),
    )

    def run():
        try:
            datum      = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_pad = config.OUTPUT_DIR / f"KING_import_{datum}.xlsx"
            json_pad   = config.OUTPUT_DIR / f"KING_import_{datum}_waarden.json"

            bouw_excel(
                producten, output_pad, client, model,
                attribuutsets, categorieen, log, progress_func,
                zoekwoorden_data=zoekwoorden_data,
                merk_domeinen=merk_domeinen,
                output_json_pad=json_pad,
                genereer_de=genereer_de,
                genereer_jl=genereer_jl,
            )

            with progress_lock:
                progress_data["output_bestand"] = output_pad.name
                progress_data["waarden_bestand"] = json_pad.name if json_pad.exists() else None
                progress_data["done"]           = True
                progress_data["running"]        = False

        except Exception as e:
            with progress_lock:
                progress_data["error"]   = str(e)
                progress_data["done"]    = True
                progress_data["running"] = False
            log(f"FOUT: {e}", "error")

    threading.Thread(target=run, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/stop", methods=["POST"])
def stop():
    with progress_lock:
        progress_data["running"] = False
    return jsonify({"ok": True})


@app.route("/progress")
def progress():
    with progress_lock:
        data = {
            "running":         progress_data["running"],
            "current":         progress_data["current"],
            "total":           progress_data["total"],
            "done":            progress_data["done"],
            "error":           progress_data["error"],
            "output_bestand":  progress_data["output_bestand"],
            "log":             progress_data["log"][-60:],
        }
    return jsonify(data)


@app.route("/download/<filename>")
def download(filename):
    pad = config.OUTPUT_DIR / filename
    if not pad.exists():
        return "Bestand niet gevonden", 404
    return send_file(pad, as_attachment=True)


@app.route("/reset", methods=["POST"])
def reset():
    with progress_lock:
        progress_data.update({
            "running": False, "current": 0, "total": 0,
            "log": [], "error": None, "done": False,
            "output_bestand": None,
        })
    for f in [config.UPLOAD_DIR / "werkdocument.json"]:
        if f.exists():
            f.unlink()
    return jsonify({"ok": True})


@app.route("/formulier")
def formulier():
    return render_template("formulier.html")


@app.route("/formulier-submit", methods=["POST"])
def formulier_submit():
    try:
        data = request.json
        if not data:
            return jsonify({"ok": False, "error": "Geen data ontvangen"})

        producten = data.get("producten", [])
        if not producten:
            return jsonify({"ok": False, "error": "Geen producten in payload"})

        for p in producten:
            if not p.get("Artikelnummer") or not p.get("Omschrijving") or not p.get("Merk"):
                return jsonify({"ok": False, "error": "Artikelnummer, Omschrijving en Merk zijn verplicht"})

        with open(config.UPLOAD_DIR / "werkdocument.json", "w", encoding="utf-8") as fp:
            json.dump(producten, fp, ensure_ascii=False, default=str)

        return jsonify({"ok": True, "count": len(producten)})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)})


@app.route("/template")
def template_download():
    # (naam, verplicht, voorbeeld, uitleg)
    KOLOMMEN = [
        ("Artikelnummer",               True,  "JOT-12345",
         "Verplicht. Het KING-artikelnummer voor dit product."),
        ("Omschrijving",                True,  "Jotun Jotaprimer 604 Grijs 5L",
         "Verplicht. Volledige productnaam inclusief variant/inhoud. Hoe meer detail, hoe beter de AI-teksten."),
        ("Merk",                        True,  "Jotun",
         "Verplicht. Merknaam zoals in KING (bijv. Jotun, Tec7, Sika)."),
        ("URL_leverancier",             False, "https://www.leverancier.nl/product/...",
         "Aanbevolen. URL van de productpagina bij de leverancier. De tool haalt automatisch de paginatekst op "
         "als basis voor de AI-omschrijving, én zoekt PDF-links voor het productinformatieblad en veiligheidsblad."),
        ("Productinformatieblad_pad",   False, r"C:\paden\pib.pdf",
         "Optioneel. Handmatig pad/URL naar het productinformatieblad. Wordt automatisch ingevuld vanuit de "
         "leverancierspagina als URL_leverancier is opgegeven en er een PIB-link gevonden wordt."),
        ("Productveiligheidsblad_pad",  False, r"C:\paden\pvb.pdf",
         "Optioneel. Handmatig pad/URL naar het productveiligheidsblad. Wordt automatisch ingevuld vanuit de "
         "leverancierspagina als URL_leverancier is opgegeven en er een VIB-link gevonden wordt."),
        ("Extra_info",                  False, "Inhoud: 300ml. Treksterkte: 3 N/mm². Temp: -40 tot +90°C.",
         "Optioneel. Extra specs of tekst die de AI moet gebruiken. Wordt gecombineerd met de paginatekst "
         "als URL_leverancier ook is ingevuld."),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Producten"

    hdr_fill    = PatternFill("solid", fgColor="002B5C")
    opt_fill    = PatternFill("solid", fgColor="00A3A1")
    tip_fill    = PatternFill("solid", fgColor="FEF9C3")
    hdr_font    = Font(bold=True, color="FFFFFF")
    ex_font     = Font(italic=True, color="6B7280")
    tip_font    = Font(italic=True, color="92400E", size=9)
    center      = Alignment(horizontal="center")
    wrap        = Alignment(wrap_text=True, vertical="top")

    for col, (naam, verplicht, voorbeeld, uitleg) in enumerate(KOLOMMEN, start=1):
        letter = openpyxl.utils.get_column_letter(col)

        # Rij 1: kolomnaam
        cel = ws.cell(row=1, column=col, value=naam)
        cel.font      = hdr_font
        cel.fill      = hdr_fill if verplicht else opt_fill
        cel.alignment = center

        # Rij 2: uitleg
        cel2 = ws.cell(row=2, column=col, value=uitleg)
        cel2.font      = tip_font
        cel2.fill      = tip_fill
        cel2.alignment = wrap

        # Rij 3: voorbeeldwaarde
        ws.cell(row=3, column=col, value=voorbeeld).font = ex_font

        ws.column_dimensions[letter].width = max(len(naam) + 4, 28)

    ws.row_dimensions[2].height = 52

    ws.cell(row=4, column=1, value="← Vul hier jouw producten in (verwijder rij 3)").font = Font(italic=True, color="B45309")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name="Nieuw_Product_Template.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/push-config")
def push_config_status():
    cfg = _king_config_from_env()
    return jsonify({
        "configured": cfg is not None,
        "host": config.KING_HOST or None,
        "administratie": config.KING_ADMINISTRATIE or None,
    })


@app.route("/push", methods=["POST"])
def push_naar_king():
    dry_run = bool((request.json or {}).get("dry_run", False))

    with push_lock:
        if push_data["running"]:
            return jsonify({"ok": False, "error": "Push al bezig"})

    cfg = _king_config_from_env()
    if not cfg and not dry_run:
        return jsonify({"ok": False, "error": "KING-verbinding niet geconfigureerd (stel KING_HOST, KING_ADMINISTRATIE, KING_ACCESS_TOKEN in)"})
    if not cfg:
        cfg = KingConfig(protocol="http", host="dry-run", port="8082",
                         administratie="dry-run", access_token="dry-run")

    with progress_lock:
        waarden_bestand = progress_data.get("waarden_bestand")

    # Fallback: zoek meest recent waarden-JSON in output/ als progress_data leeg is
    if not waarden_bestand:
        kandidaten = sorted(config.OUTPUT_DIR.glob("*_waarden.json"), key=lambda p: p.stat().st_mtime, reverse=True)
        if kandidaten:
            waarden_bestand = kandidaten[0].name

    if not waarden_bestand:
        return jsonify({"ok": False, "error": "Genereer eerst producten via stap 3 — geen gegenereerde data gevonden"})

    waarden_pad = config.OUTPUT_DIR / waarden_bestand
    if not waarden_pad.exists():
        return jsonify({"ok": False, "error": f"Waarden-bestand niet gevonden: {waarden_bestand}"})

    try:
        with open(waarden_pad, encoding="utf-8") as f:
            alle_waarden = json.load(f)
    except Exception as e:
        return jsonify({"ok": False, "error": f"Waarden laden mislukt: {e}"})

    with push_lock:
        push_data.update({
            "running": True, "current": 0, "total": len(alle_waarden),
            "log": [], "error": None, "done": False,
            "sessie_id": None, "n_gelukt": 0, "n_mislukt": 0,
            "dry_run": dry_run,
        })

    def push_log(msg: str, typ: str = "info"):
        with push_lock:
            push_data["log"].append({
                "msg": msg, "type": typ,
                "time": datetime.now().strftime("%H:%M:%S"),
            })

    def push_progress(huidig: int, totaal: int):
        with push_lock:
            push_data["current"] = huidig
            push_data["total"]   = totaal

    def run_push():
        try:
            db  = _get_push_db()
            gebruiker = huidig_gebruiker()
            resultaat = push_batch(
                alle_waarden, cfg, db, gebruiker,
                log_func=push_log, progress_func=push_progress,
                dry_run=dry_run,
            )
            with push_lock:
                push_data.update({
                    "running":   False,
                    "done":      True,
                    "sessie_id": resultaat["sessie_id"],
                    "n_gelukt":  resultaat["n_gelukt"],
                    "n_mislukt": resultaat["n_mislukt"],
                })
        except Exception as e:
            with push_lock:
                push_data.update({
                    "running": False, "done": True, "error": str(e),
                })
            push_log(f"FOUT: {e}", "error")

    threading.Thread(target=run_push, daemon=True).start()
    return jsonify({"ok": True})


@app.route("/push-status")
def push_status():
    with push_lock:
        data = dict(push_data)
        data["log"] = data["log"][-60:]
    return jsonify(data)


@app.route("/push-geschiedenis")
def push_geschiedenis_json():
    try:
        db = _get_push_db()
        sessie_id = request.args.get("sessie_id", type=int)
        artikel_id = request.args.get("artikel_id", type=int)

        if artikel_id:
            return jsonify(db.haal_calls(artikel_id))
        if sessie_id:
            return jsonify(db.haal_artikelen(sessie_id))
        return jsonify(db.haal_sessies())
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/push-geschiedenis-html")
def push_geschiedenis_html():
    try:
        db = _get_push_db()
        sessies = db.haal_sessies(limit=50)
    except Exception:
        sessies = []
    return render_template("push_geschiedenis.html", sessies=sessies)


if __name__ == "__main__":
    import webbrowser

    def open_browser():
        time.sleep(1.5)
        webbrowser.open("http://127.0.0.1:5002")

    print("=" * 60)
    print("  Product Import Generator")
    print("  http://127.0.0.1:5002")
    print("=" * 60)
    threading.Thread(target=open_browser, daemon=True).start()
    app.run(debug=False, threaded=True, port=5002, host="127.0.0.1")

